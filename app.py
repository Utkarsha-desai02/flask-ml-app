from flask import Flask, request, redirect, url_for, session, render_template_string, send_from_directory
import matplotlib
matplotlib.use('Agg')
import os
import rasterio
import matplotlib.pyplot as plt
from matplotlib.path import Path
from matplotlib.patches import PathPatch
import io
import base64
import numpy as np
import sqlite3
import hashlib
import secrets
from datetime import datetime
from functools import wraps

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

try:
    import geopandas as gpd
    GEOPANDAS_OK = True
except ImportError:
    GEOPANDAS_OK = False

app = Flask(__name__)
app.secret_key = secrets.token_hex(32)

taluks = [
    "Bardez", "Bicholim", "Canacona", "Dharbandora",
    "Mormugoa", "Pernem", "Ponda", "Quepem",
    "Salcete", "Sanguem", "Satari", "Tiswadi"
]
models = ["CNN", "RF", "SVM", "KNN", "LR"]

MODEL_FULL_NAMES = {
    "CNN": "CNN — Deep Learning",
    "RF":  "RF — Random Forest",
    "SVM": "SVM — Support Vector Machine",
    "KNN": "KNN — K-Nearest Neighbours",
    "LR":  "LR — Logistic Regression",
}
MODEL_ICONS = {"CNN": "🧠", "RF": "🌲", "SVM": "📐", "KNN": "🔍", "LR": "📈"}

FILENAME_MAP = {
    ("Bardez","CNN"): "Bardez_CNN_Variability.tif",
    ("Bicholim","CNN"): "Bicholim_CNN_Variability.tif",
    ("Canacona","CNN"): "Canacona_CNN_Variability.tif",
    ("Dharbandora","CNN"): "Dharbandora_CNN_Variability.tif",
    ("Mormugoa","CNN"): "Mormugoa_CNN_Variability.tif",
    ("Pernem","CNN"): "Pernem_CNN_Variability.tif",
    ("Ponda","CNN"): "Ponda_CNN_Variability.tif",
    ("Quepem","CNN"): "Quepem_CNN_Variability.tif",
    ("Salcete","CNN"): "Salcete_CNN_Variability.tif",
    ("Sanguem","CNN"): "Sanguem_CNN_Variability.tif",
    ("Satari","CNN"): "Satari_CNN_Variability.tif",
    ("Tiswadi","CNN"): "Tiswadi_CNN_Variability.tif",
    ("Bardez","RF"): "Bardez_RF_Variability.tif",
    ("Bicholim","RF"): "Bicholim_RF_Variability.tif",
    ("Canacona","RF"): "canacona_RF_Variability.tif",
    ("Dharbandora","RF"): "Dharbandora_RF_Variability.tif",
    ("Mormugoa","RF"): "Mormugoa_RF_Variability.tif",
    ("Pernem","RF"): "Pernem_RF_Variability.tif",
    ("Ponda","RF"): "Ponda_RF_Variability.tif",
    ("Quepem","RF"): "Quepem_RF_Variability.tif",
    ("Salcete","RF"): "Salcete_RF_Variability.tif",
    ("Sanguem","RF"): "Sanguem_RF_Variability.tif",
    ("Satari","RF"): "satari_RF_Variability.tif",
    ("Tiswadi","RF"): "Tiswadi_RF_Variability.tif",
    ("Bardez","SVM"): "Bardez_SVM_Variability.tif",
    ("Bicholim","SVM"): "Bicholim_Vulnerability_SVM_Final.tif",
    ("Canacona","SVM"): "Canacona_SVM_Variability.tif",
    ("Dharbandora","SVM"): "Dharbandora_SVM_Variability.tif",
    ("Mormugoa","SVM"): "Mormugoa_SVM_Variability.tif",
    ("Pernem","SVM"): "Pernem_SVM_Variability.tif",
    ("Ponda","SVM"): "Ponda_SVM_Variability.tif",
    ("Quepem","SVM"): "Quepem_SVM_Variability.tif",
    ("Salcete","SVM"): "Salcete_SVM_Variability.tif",
    ("Sanguem","SVM"): "Sanguem_Vulnerability_SVM_Final.tif",
    ("Satari","SVM"): "Satari_SVM_Variability.tif",
    ("Tiswadi","SVM"): "Tiswadi_SVM_Variability.tif",
    ("Bardez","KNN"): "Bardez_KNN_Variability.tif",
    ("Bicholim","KNN"): "Bicholim_KNN_Variability.tif",
    ("Canacona","KNN"): "Canacona_KNN_Variability.tif",
    ("Dharbandora","KNN"): "Dharbandora_KNN_Variability.tif",
    ("Mormugoa","KNN"): "Mormugoa_KNN_Variability.tif",
    ("Pernem","KNN"): "Pernem_KNN_Variability.tif",
    ("Ponda","KNN"): "Ponda_KNN_Variability.tif",
    ("Quepem","KNN"): "Quepem_KNN_Variability.tif",
    ("Salcete","KNN"): "Salcete_KNN_Variability.tif",
    ("Sanguem","KNN"): "Sanguem_KNN_Variability.tif",
    ("Satari","KNN"): "Satari_KNN_Variability.tif",
    ("Tiswadi","KNN"): "Tiswadi_KNN_Variability.tif",
    ("Bardez","LR"): "Bardez_LR_Variability.tif",
    ("Bicholim","LR"): "Bicholim_LR_Variability.tif",
    ("Canacona","LR"): "Canacona_LR_Variability.tif",
    ("Dharbandora","LR"): "Dharbandora_LR_Variability.tif",
    ("Mormugoa","LR"): "Mormugoa_LR_Variability.tif",
    ("Pernem","LR"): "Pernem_LR_Variability.tif",
    ("Ponda","LR"): "Ponda_LR_Variability.tif",
    ("Quepem","LR"): "Quepem_LR_Variability.tif",
    ("Salcete","LR"): "Salcete_LR_Variability.tif",
    ("Sanguem","LR"): "Sanguem_LR_Variability.tif",
    ("Satari","LR"): "Satari_LR_Variability.tif",
    ("Tiswadi","LR"): "Tiswadi_LR_Variability.tif",
}

BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
GOA_ML_FOLDER = "ML Models"

GOA_OVERVIEW_FILES = {
    "CNN": ("CNN",                  "Landslide_Vulnerability_CNN_Final10.tif"),
    "RF":  ("Random forest",        "Landslide_Probability_RF_Final99.tif"),
    "SVM": ("SVM model",            "Landslide_Vulnerability_SVM_Final.tif"),
    "KNN": ("KNN model",            "Landslide_Vulnerability_KNN(1).tif"),
    "LR":  ("Logistic Rregression", "Landslide_Vulnerability_LR_Final.tif"),
}

EXCEL_PATH            = os.path.join(BASE_DIR, "user_registrations.xlsx")
DB_PATH               = os.path.join(BASE_DIR, "eva_system.db")
TALUKA_SHAPEFILE_PATH = os.path.join(BASE_DIR, "static", "data", "Goa _talukas.shp")
TALUKA_SHAPEFILE_ALT  = os.path.join(BASE_DIR, "static", "data", "Goa _talukas.shp")

TALUKA_CENTROIDS = {
    "Pernem":      (73.790, 15.720),
    "Bardez":      (73.760, 15.555),
    "Bicholim":    (73.945, 15.600),
    "Satari":      (74.085, 15.690),
    "Tiswadi":     (73.870, 15.490),
    "Ponda":       (74.000, 15.400),
    "Dharbandora": (74.220, 15.335),
    "Mormugoa":    (73.820, 15.390),
    "Salcete":     (73.930, 15.195),
    "Quepem":      (74.070, 15.215),
    "Sanguem":     (74.230, 15.100),
    "Canacona":    (74.020, 14.990),
}

SHAPEFILE_NAME_MAP = {
    "Bardez": "Bardez", "Bicholim": "Bicholim", "Canacona": "Canacona",
    "Dharbandora": "Dharbandora", "Mormugoa": "Mormugoa", "Pernem": "Pernem",
    "Ponda": "Ponda", "Quepem": "Quepem", "Salcete": "Salcete",
    "Sanguem": "Sanguem", "Satari": "Satari", "Tiswadi": "Tiswadi",
}

from matplotlib.colors import LinearSegmentedColormap as _LSC
RISK_CMAP = _LSC.from_list("risk", [
    "#1a9850", "#66bd63", "#a6d96a", "#fee08b",
    "#fdae61", "#f46d43", "#d73027", "#a50026",
], N=256)

_gdf_cache = None

def get_gdf():
    global _gdf_cache
    if _gdf_cache is not None:
        return _gdf_cache
    if not GEOPANDAS_OK:
        return None
    shp = None
    for p in [TALUKA_SHAPEFILE_PATH, TALUKA_SHAPEFILE_ALT]:
        if os.path.exists(p):
            shp = p; break
    if not shp:
        print("[GDF] No shapefile found.")
        return None
    try:
        gdf = gpd.read_file(shp)
        if gdf.crs and gdf.crs.to_epsg() != 4326:
            gdf = gdf.to_crs(epsg=4326)
        _gdf_cache = gdf
        print(f"[GDF] Loaded {len(gdf)} features | cols: {gdf.columns.tolist()}")
    except Exception as e:
        print(f"[GDF] Error: {e}")
    return _gdf_cache

def gdf_name_col(gdf):
    for c in ['NAME_3', 'NAME_2', 'TALUKA', 'name', 'NAME']:
        if c in gdf.columns:
            return c
    return gdf.columns[0]

def get_goa_file_path(model):
    if model not in GOA_OVERVIEW_FILES:
        return None, None
    subfolder, fname = GOA_OVERVIEW_FILES[model]
    full_path = os.path.join(BASE_DIR, GOA_ML_FOLDER, subfolder, fname)
    if not os.path.exists(full_path):
        folder = os.path.join(BASE_DIR, GOA_ML_FOLDER, subfolder)
        if os.path.isdir(folder):
            for f in sorted(os.listdir(folder)):
                if "vulnerability" in f.lower() and f.lower().endswith(".tif"):
                    return os.path.join(folder, f), f
    return full_path, fname

def get_file_path(taluk, model):
    fname = FILENAME_MAP.get((taluk, model))
    if fname:
        p = os.path.join(BASE_DIR, "tif files", model, fname)
        if os.path.exists(p):
            return p, fname
    folder = os.path.join(BASE_DIR, "tif files", model)
    if os.path.isdir(folder):
        tk = taluk.lower().replace(" ", "")
        mk = model.lower()
        for f in sorted(os.listdir(folder)):
            fl = f.lower().replace(" ", "")
            if tk in fl and mk in fl and fl.endswith(".tif"):
                return os.path.join(folder, f), f
    fallback = fname or f"{taluk}_{model}_Variability.tif"
    return os.path.join(BASE_DIR, "tif files", model, fallback), fallback

def init_excel():
    if not EXCEL_OK or os.path.exists(EXCEL_PATH):
        return
    wb = Workbook(); ws = wb.active; ws.title = "User Registrations"
    headers = ["#", "Username", "Email", "Role", "Registered At"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="0D9488")
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")
    for col in range(1, 6):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = [6,22,32,12,22][col-1]
    wb.save(EXCEL_PATH)

def append_user_to_excel(username, email, role, registered_at):
    if not EXCEL_OK:
        return
    try:
        if not os.path.exists(EXCEL_PATH):
            init_excel()
        wb = load_workbook(EXCEL_PATH); ws = wb.active
        nr = ws.max_row + 1; num = nr - 1
        fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="F0FDFA" if num%2==0 else "FFFFFF")
        for col, val in enumerate([num, username, email, role, registered_at], 1):
            cell = ws.cell(row=nr, column=col, value=val)
            cell.fill = fill; cell.alignment = openpyxl.styles.Alignment(horizontal="center")
        wb.save(EXCEL_PATH)
    except Exception as e:
        print(f"[Excel] {e}")

def get_db():
    conn = sqlite3.connect(DB_PATH, timeout=30, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=30000")
    return conn

def hash_password(p):
    return hashlib.sha256(p.encode()).hexdigest()

def init_db():
    conn = get_db(); c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL, email TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL, role TEXT NOT NULL DEFAULT 'user',
        created_at TEXT NOT NULL)""")
    c.execute("""CREATE TABLE IF NOT EXISTS assessment_logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER, username TEXT, taluk TEXT, model TEXT, timestamp TEXT,
        FOREIGN KEY (user_id) REFERENCES users(id))""")
    try:
        now = datetime.now().isoformat()
        c.execute("INSERT INTO users (username,email,password_hash,role,created_at) VALUES (?,?,?,?,?)",
                  ("admin","admin@eva-goa.gov.in",hash_password("admin123"),"admin",now))
        conn.commit()
        append_user_to_excel("admin","admin@eva-goa.gov.in","admin",now[:16])
    except sqlite3.IntegrityError:
        pass
    conn.commit(); conn.close()

def log_assessment(user_id, username, taluk, model):
    conn = get_db()
    try:
        conn.execute("INSERT INTO assessment_logs (user_id,username,taluk,model,timestamp) VALUES (?,?,?,?,?)",
                     (user_id, username, taluk, model, datetime.now().isoformat()))
        conn.commit()
    finally:
        conn.close()

def login_required(f):
    @wraps(f)
    def d(*a, **kw):
        if 'user_id' not in session: return redirect(url_for('login'))
        return f(*a, **kw)
    return d

def admin_required(f):
    @wraps(f)
    def d(*a, **kw):
        if 'user_id' not in session: return redirect(url_for('login'))
        if session.get('role') != 'admin': return redirect(url_for('dashboard'))
        return f(*a, **kw)
    return d

@app.route('/static/images/<path:filename>')
def serve_static_image(filename):
    img_dir = os.path.join(BASE_DIR, 'static', 'images')
    return send_from_directory(img_dir, filename)


# =============================================================================
#  LANDING PAGE — Goa on India map, project name updated, no AI words
# =============================================================================
LANDING_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>ML-Based Environmental Vulnerability Assessment for Goa</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Serif+Display:ital@0;1&family=DM+Sans:wght@300;400;500;600&family=Space+Mono:wght@400;700&display=swap" rel="stylesheet">
<style>
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{
  --teal:#0d9488;--emerald:#059669;
  --ink:#0a1410;--mid:#3d5450;--soft:#7a8f8c;
  --paper:#f7faf9;--border:rgba(13,148,136,0.15);
}
html{scroll-behavior:smooth}
body{font-family:'DM Sans',sans-serif;background:var(--paper);color:var(--ink);overflow-x:hidden;}

/* NAV */
nav{position:fixed;top:0;left:0;right:0;z-index:1000;display:flex;align-items:center;justify-content:space-between;padding:18px 56px;transition:all 0.4s;}
nav.scrolled{padding:12px 56px;background:rgba(247,250,249,0.97);backdrop-filter:blur(20px);border-bottom:1px solid var(--border);box-shadow:0 4px 20px rgba(10,20,16,0.06);}
.nav-logo{display:flex;align-items:center;gap:11px;text-decoration:none;}
.nav-logo-mark{width:36px;height:36px;background:linear-gradient(135deg,var(--teal),var(--emerald));border-radius:9px;display:flex;align-items:center;justify-content:center;}
.nav-logo-mark img{width:28px;height:28px;object-fit:contain;border-radius:5px;}
.nav-brand{font-family:'DM Serif Display',serif;font-size:13px;color:var(--ink);max-width:230px;line-height:1.3;}
.nav-brand span{color:var(--teal);}
.nav-links{display:flex;gap:28px;}
.nav-links a{color:var(--mid);text-decoration:none;font-size:13px;font-weight:500;transition:color 0.2s;}
.nav-links a:hover{color:var(--teal);}
.nav-auth{display:flex;gap:10px;}
.nb1{padding:8px 17px;border:1.5px solid var(--border);border-radius:50px;font-size:13px;color:var(--mid);text-decoration:none;transition:all 0.2s;}
.nb1:hover{border-color:var(--teal);color:var(--teal);}
.nb2{padding:8px 18px;background:var(--ink);border-radius:50px;font-size:13px;color:#fff;text-decoration:none;transition:all 0.2s;}
.nb2:hover{background:var(--teal);}

/* ══ HERO — goa.png FULL BACKGROUND ══ */
#hero{
  position:relative;
  width:100vw;
  min-height:100vh;
  display:flex;
  align-items:center;
  overflow:hidden;
}
.hero-bg{
  position:absolute;
  inset:0;
  background-image:url('/static/images/goa.png');
  background-size:50;
 background-position:right 8% 40%;
  background-repeat:no-repeat;
  z-index:0;
}
.hero-body{
  position:relative;
  z-index:2;
  padding:140px 80px 80px;
  max-width:680px;
}
.badge{display:inline-flex;align-items:center;gap:8px;padding:6px 16px;background:rgba(13,148,136,0.10);border:1px solid rgba(13,148,136,0.22);border-radius:50px;margin-bottom:24px;font-family:'Space Mono',monospace;font-size:9px;letter-spacing:2px;color:var(--teal);text-transform:uppercase;}
.badge-dot{width:5px;height:5px;border-radius:50%;background:var(--teal);animation:blink 1.8s infinite;}
@keyframes blink{0%,100%{opacity:1}50%{opacity:0.3}}
h1.hero-title{font-family:'DM Serif Display',serif;font-size:clamp(34px,4.2vw,60px);line-height:1.08;letter-spacing:-1.5px;color:var(--ink);margin-bottom:20px;}
h1.hero-title em{font-style:italic;color:var(--teal);}
.hero-sub{font-size:16px;line-height:1.85;color:var(--mid);max-width:500px;margin-bottom:36px;}

/* Buttons */
.hero-btns{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:28px;}
.btn-a{display:inline-flex;align-items:center;gap:8px;padding:13px 26px;background:linear-gradient(135deg,var(--teal),var(--emerald));color:#fff;border-radius:50px;font-size:14px;font-weight:600;text-decoration:none;box-shadow:0 8px 22px rgba(13,148,136,0.28);transition:all 0.3s;}
.btn-a:hover{transform:translateY(-2px);box-shadow:0 12px 30px rgba(13,148,136,0.38);}
.btn-b{display:inline-flex;align-items:center;gap:8px;padding:12px 22px;background:#1e3a5f;color:#fff;border-radius:50px;font-size:14px;font-weight:600;text-decoration:none;box-shadow:0 8px 20px rgba(30,58,95,0.2);transition:all 0.3s;}
.btn-b:hover{transform:translateY(-2px);background:#152f50;}
.btn-c{display:inline-flex;align-items:center;gap:8px;padding:12px 20px;color:var(--mid);border:1.5px solid rgba(61,84,80,0.2);border-radius:50px;font-size:14px;font-weight:500;text-decoration:none;background:rgba(255,255,255,0.75);backdrop-filter:blur(6px);transition:all 0.3s;}
.btn-c:hover{border-color:var(--teal);color:var(--teal);}

/* Access role cards */
.role-row{display:flex;gap:11px;flex-wrap:wrap;margin-bottom:36px;}
.rc{display:flex;align-items:center;gap:10px;padding:13px 16px;border-radius:15px;background:rgba(255,255,255,0.88);border:1.5px solid var(--border);text-decoration:none;min-width:150px;transition:all 0.3s;backdrop-filter:blur(10px);}
.rc:hover{transform:translateY(-3px);box-shadow:0 10px 28px rgba(13,148,136,0.12);border-color:rgba(13,148,136,0.28);}
.rc.rc-admin{border-color:rgba(30,58,95,0.16);background:rgba(240,244,255,0.88);}
.rc.rc-admin:hover{box-shadow:0 10px 28px rgba(30,58,95,0.10);border-color:rgba(30,58,95,0.3);}
.rc-icon{font-size:19px;}
.rc-title{font-size:13px;font-weight:600;color:var(--ink);display:block;}
.rc-sub{font-family:'Space Mono',monospace;font-size:8px;letter-spacing:1.5px;text-transform:uppercase;color:var(--teal);}
.rc-admin .rc-sub{color:#1e3a5f;}

/* Stats */
.hero-stats{display:flex;gap:28px;padding-top:24px;border-top:1px solid var(--border);}
.hs-num{font-family:'DM Serif Display',serif;font-size:30px;color:var(--ink);letter-spacing:-1px;display:block;}
.hs-num span{color:var(--teal);}
.hs-lbl{font-family:'Space Mono',monospace;font-size:8px;letter-spacing:1.5px;text-transform:uppercase;color:var(--soft);}

/* TICKER */
.ticker-wrap{overflow:hidden;padding:12px 0;border-top:1px solid var(--border);border-bottom:1px solid var(--border);}
.ticker{display:flex;gap:16px;animation:tick 24s linear infinite;white-space:nowrap;}
@keyframes tick{0%{transform:translateX(0)}100%{transform:translateX(-50%)}}
.tc{display:inline-flex;align-items:center;gap:6px;padding:6px 15px;border:1px solid var(--border);border-radius:50px;font-family:'Space Mono',monospace;font-size:9px;letter-spacing:2px;color:var(--mid);text-transform:uppercase;background:rgba(255,255,255,0.6);flex-shrink:0;}
.tc::before{content:'';width:4px;height:4px;border-radius:50%;background:var(--teal);}

/* STATS ROW */
.stats-row{display:flex;border-top:1.5px solid var(--border);border-bottom:1.5px solid var(--border);}
.sc{flex:1;padding:40px 32px;border-right:1.5px solid var(--border);text-align:center;opacity:0;transform:translateY(16px);transition:opacity 0.6s,transform 0.6s;}
.sc:last-child{border-right:none;}
.sc.vis{opacity:1;transform:translateY(0);}
.sc-num{font-family:'DM Serif Display',serif;font-size:48px;color:var(--ink);letter-spacing:-2px;display:block;margin-bottom:5px;}
.sc-num span{color:var(--teal);}
.sc-lbl{font-family:'Space Mono',monospace;font-size:9px;letter-spacing:2px;color:var(--soft);text-transform:uppercase;}

/* SECTIONS */
.section{padding:90px 80px;}
.section-tinted{background:linear-gradient(135deg,#edfaf7,#f5fffe);}
.si{max-width:1180px;margin:0 auto;display:grid;grid-template-columns:1fr 1fr;gap:72px;align-items:center;}
.si.rev{direction:rtl;}
.si.rev>*{direction:ltr;}
.stag{display:inline-block;padding:5px 13px;background:rgba(13,148,136,0.1);border:1px solid rgba(13,148,136,0.2);border-radius:50px;font-family:'Space Mono',monospace;font-size:9px;letter-spacing:2px;color:var(--teal);text-transform:uppercase;margin-bottom:16px;}
.snum{font-family:'Space Mono',monospace;font-size:9px;letter-spacing:3px;color:var(--soft);text-transform:uppercase;margin-bottom:12px;}
.stitle{font-family:'DM Serif Display',serif;font-size:clamp(28px,3.2vw,46px);line-height:1.1;letter-spacing:-1.5px;color:var(--ink);margin-bottom:14px;}
.stitle em{font-style:italic;color:var(--teal);}
.sbody{font-size:15px;line-height:1.85;color:var(--mid);}

.card{background:rgba(255,255,255,0.92);border:1.5px solid var(--border);border-radius:20px;padding:26px;box-shadow:0 14px 44px rgba(13,148,136,0.08);position:relative;overflow:hidden;}
.card::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;background:linear-gradient(90deg,var(--teal),var(--emerald));}
.card-lbl{font-family:'Space Mono',monospace;font-size:9px;letter-spacing:2px;color:var(--soft);text-transform:uppercase;margin-bottom:16px;}
.frow{display:flex;align-items:center;gap:10px;margin-bottom:11px;}
.fl{font-family:'Space Mono',monospace;font-size:9px;color:var(--mid);width:100px;flex-shrink:0;text-transform:uppercase;}
.fb{flex:1;height:3px;background:rgba(13,148,136,0.1);border-radius:2px;overflow:hidden;}
.ff{height:100%;background:linear-gradient(90deg,var(--teal),var(--emerald));width:0;transition:width 1.2s cubic-bezier(0.16,1,0.3,1);}
.fv{font-family:'Space Mono',monospace;font-size:9px;color:var(--teal);width:32px;text-align:right;}
.mrow{display:flex;align-items:center;justify-content:space-between;padding:11px 0;border-bottom:1px solid var(--border);}
.mrow:last-child{border-bottom:none;}
.ml{display:flex;align-items:center;gap:10px;}
.mn{font-size:13px;font-weight:600;color:var(--ink);}
.mf{font-family:'Space Mono',monospace;font-size:9px;color:var(--soft);}
.ms{font-family:'Space Mono',monospace;font-size:9px;color:var(--teal);}
.cgrid{display:grid;grid-template-columns:1fr 1fr;gap:8px;}
.ci{display:flex;align-items:center;gap:7px;padding:9px 11px;background:rgba(13,148,136,0.04);border:1px solid var(--border);border-radius:9px;}
.cd{width:5px;height:5px;border-radius:50%;background:var(--teal);}
.cn{font-family:'Space Mono',monospace;font-size:9px;letter-spacing:1.5px;color:var(--mid);text-transform:uppercase;}

/* FEATURES */
.fg{display:grid;grid-template-columns:1fr 1fr;gap:18px;max-width:1180px;margin:0 auto;padding:0 80px 80px;}
.fi{padding:28px;border-radius:18px;background:rgba(255,255,255,0.85);border:1.5px solid var(--border);transition:all 0.3s;opacity:0;transform:translateY(18px);}
.fi.vis{opacity:1;transform:translateY(0);}
.fi:hover{transform:translateY(-4px);box-shadow:0 14px 44px rgba(13,148,136,0.11);}
.fi-icon{width:46px;height:46px;border-radius:12px;background:rgba(13,148,136,0.1);display:flex;align-items:center;justify-content:center;font-size:20px;margin-bottom:14px;}
.fi-title{font-family:'DM Serif Display',serif;font-size:19px;color:var(--ink);margin-bottom:8px;}
.fi-body{font-size:14px;line-height:1.75;color:var(--mid);}

/* CTA */
.cta{padding:80px;text-align:center;background:linear-gradient(135deg,#edfaf7,var(--paper));border-top:1.5px solid var(--border);}
.cta-title{font-family:'DM Serif Display',serif;font-size:clamp(32px,4.5vw,58px);color:var(--ink);letter-spacing:-2px;margin-bottom:14px;line-height:1.06;}
.cta-title em{font-style:italic;color:var(--teal);}
.cta-sub{font-size:16px;color:var(--mid);line-height:1.8;max-width:440px;margin:0 auto 36px;}
.cta-cards{display:flex;justify-content:center;gap:16px;flex-wrap:wrap;}
.cc{display:flex;flex-direction:column;align-items:center;gap:8px;padding:24px 20px;border-radius:18px;text-decoration:none;min-width:170px;position:relative;overflow:hidden;transition:all 0.3s;}
.cc::before{content:'';position:absolute;top:0;left:0;right:0;height:3px;}
.cc-r{background:rgba(13,148,136,0.07);border:1.5px solid rgba(13,148,136,0.22);}
.cc-r::before{background:linear-gradient(90deg,var(--teal),var(--emerald));}
.cc-r:hover{transform:translateY(-4px);box-shadow:0 12px 32px rgba(13,148,136,0.14);}
.cc-a{background:rgba(30,58,95,0.06);border:1.5px solid rgba(30,58,95,0.18);}
.cc-a::before{background:linear-gradient(90deg,#1e3a5f,#2563eb);}
.cc-a:hover{transform:translateY(-4px);box-shadow:0 12px 32px rgba(30,58,95,0.11);}
.cc-icon{font-size:28px;}
.cc-title{font-family:'DM Serif Display',serif;font-size:16px;color:var(--ink);}
.cc-desc{font-family:'Space Mono',monospace;font-size:8px;letter-spacing:1.5px;text-transform:uppercase;color:var(--soft);text-align:center;}
.cc-btn{padding:6px 16px;border-radius:50px;font-size:11px;font-weight:600;color:#fff;margin-top:3px;}
.cb-r{background:linear-gradient(135deg,var(--teal),var(--emerald));}
.cb-a{background:linear-gradient(135deg,#1e3a5f,#2563eb);}

footer{padding:24px 80px;display:flex;align-items:center;justify-content:space-between;border-top:1px solid var(--border);font-family:'Space Mono',monospace;font-size:9px;letter-spacing:1.5px;color:var(--soft);text-transform:uppercase;}

@media(max-width:900px){
  .hero-body{padding:100px 24px 60px;max-width:100%;}
  .si{grid-template-columns:1fr;gap:32px;padding:0;}
  .section{padding:55px 24px;}
  .fg{grid-template-columns:1fr;padding:0 24px 55px;}
  .stats-row,.cta-cards{flex-direction:column;}
  .sc{border-right:none;border-bottom:1.5px solid var(--border);}
  nav{padding:13px 20px;}
  footer{flex-direction:column;gap:8px;text-align:center;padding:18px;}
  .cta{padding:55px 24px;}
}
</style>
</head>
<body>

<nav id="nav">
  <a class="nav-logo" href="/">
    <div class="nav-logo-mark">
      <img src="/static/images/logo.jpg" alt="EVA" onerror="this.parentElement.innerHTML='🌿'">
    </div>
    <div class="nav-brand">ML-Based Environmental<br><span>Vulnerability Assessment · Goa</span></div>
  </a>
  <div class="nav-links">
    <a href="#how">How It Works</a>
    <a href="#coverage">Areas</a>
    <a href="#models">Models</a>
  </div>
  <div class="nav-auth">
    <a href="{{ url_for('login') }}" class="nb1">Researcher Login</a>
    <a href="{{ url_for('authority_login') }}" class="nb2">Authority Login</a>
  </div>
</nav>

<!-- ═══════════ HERO — GOA.PNG FULL BACKGROUND, NO SVG ═══════════ -->
<section id="hero">
  <div class="hero-bg"></div>
  <div class="hero-overlay"></div>
  <div class="hero-body">
    <div class="badge"><div class="badge-dot"></div>Final Year Research Project</div>
    <h1 class="hero-title">
      Machine Learning Based<br>
      <em>Environmental Vulnerability</em><br>
      Assessment for Goa
    </h1>
    <p class="hero-sub">Landslide risk mapped across all 12 talukas of Goa using five computer models trained on satellite images — helping researchers and authorities understand where the land is most at risk.</p>

    <div class="hero-btns">
      <a href="{{ url_for('register') }}" class="btn-a">👤 Researcher Sign Up</a>
      <a href="{{ url_for('authority_register') }}" class="btn-b">🏛️ Authority Register</a>
      <a href="{{ url_for('login') }}" class="btn-c">Sign In</a>
    </div>

    <div class="role-row">
      <a href="{{ url_for('login') }}" class="rc">
        <span class="rc-icon">👤</span>
        <div><span class="rc-title">Researcher</span><span class="rc-sub">Login →</span></div>
      </a>
      <a href="{{ url_for('register') }}" class="rc">
        <span class="rc-icon">📝</span>
        <div><span class="rc-title">Register</span><span class="rc-sub">New Account →</span></div>
      </a>
      <a href="{{ url_for('authority_login') }}" class="rc rc-admin">
        <span class="rc-icon">🏛️</span>
        <div><span class="rc-title">Authority</span><span class="rc-sub">Login →</span></div>
      </a>
      <a href="{{ url_for('authority_register') }}" class="rc rc-admin">
        <span class="rc-icon">🔐</span>
        <div><span class="rc-title">Auth. Register</span><span class="rc-sub">Apply →</span></div>
      </a>
    </div>

    <div class="hero-stats">
      <div><span class="hs-num">12</span><span class="hs-lbl">Talukas</span></div>
      <div><span class="hs-num">5</span><span class="hs-lbl">Models</span></div>
      <div><span class="hs-num">60<span>+</span></span><span class="hs-lbl">Risk Maps</span></div>
    </div>
  </div>
</section>
<!-- ═══════════════════════════════════════════════════════════════ -->

<div class="ticker-wrap">
  <div class="ticker">
    {% for t in taluks %}<div class="tc">{{ t }}</div>{% endfor %}
    {% for t in taluks %}<div class="tc">{{ t }}</div>{% endfor %}
    {% for t in taluks %}<div class="tc">{{ t }}</div>{% endfor %}
  </div>
</div>

<div class="stats-row">
  <div class="sc" data-delay="0"><span class="sc-num" id="cnt1">0</span><div class="sc-lbl">Areas fully mapped</div></div>
  <div class="sc" data-delay="100"><span class="sc-num" id="cnt2">0</span><div class="sc-lbl">Models deployed</div></div>
  <div class="sc" data-delay="200"><span class="sc-num"><span id="cnt3">0</span><span style="color:var(--teal)">+</span></span><div class="sc-lbl">Risk maps available</div></div>
  <div class="sc" data-delay="300"><span class="sc-num" style="font-size:36px">Goa</span><div class="sc-lbl">Complete state coverage</div></div>
</div>

<section class="section" id="how">
  <div class="si">
    <div data-reveal>
      <div class="snum">01 / 04</div>
      <div class="stag">How It Works</div>
      <h2 class="stitle">Goa's landslide risk,<br>mapped <em>area by area</em></h2>
      <p class="sbody">Satellite images of Goa were used to train models to recognise high-risk terrain. Key factors include vegetation, water presence, soil moisture, land use, and historical landslide data.</p>
    </div>
    <div data-reveal>
      <div class="card">
        <div class="card-lbl">Risk Factors Studied</div>
        <div class="frow"><span class="fl">Vegetation</span><div class="fb"><div class="ff" data-w="82"></div></div><span class="fv">82%</span></div>
        <div class="frow"><span class="fl">Water Cover</span><div class="fb"><div class="ff" data-w="70"></div></div><span class="fv">70%</span></div>
        <div class="frow"><span class="fl">Soil Moisture</span><div class="fb"><div class="ff" data-w="65"></div></div><span class="fv">65%</span></div>
        <div class="frow"><span class="fl">Land Use</span><div class="fb"><div class="ff" data-w="75"></div></div><span class="fv">75%</span></div>
        <div class="frow"><span class="fl">Slide History</span><div class="fb"><div class="ff" data-w="88"></div></div><span class="fv">88%</span></div>
      </div>
    </div>
  </div>
</section>

<section class="section section-tinted" id="models">
  <div class="si rev">
    <div data-reveal>
      <div class="snum">02 / 04</div>
      <div class="stag">Models</div>
      <h2 class="stitle">Five models,<br><em></em></h2>
      <p class="sbody">CNN, Random Forest, SVM, K-Nearest Neighbours, and Logistic Regression each analyse the same satellite data differently. Agreement across models means high confidence.</p>
    </div>
    <div data-reveal>
      <div class="card">
        <div class="card-lbl">Available Models</div>
        {% for m, name in model_names.items() %}
        <div class="mrow">
          <div class="ml">
            <span style="font-size:16px">{{ icons[m] }}</span>
            <div><div class="mn">{{ m }}</div><div class="mf">{{ name.split('—')[1].strip() if '—' in name else name }}</div></div>
          </div>
          <span class="ms">● Active</span>
        </div>
        {% endfor %}
      </div>
    </div>
  </div>
</section>

<section class="section" id="coverage">
  <div class="si">
    <div data-reveal>
      <div class="snum">03 / 04</div>
      <div class="stag">Coverage</div>
      <h2 class="stitle">All 12 talukas,<br><em>fully covered</em></h2>
      <p class="sbody">From Pernem in the north to Canacona in the south — every area of Goa has a complete risk assessment based on actual satellite data.</p>
    </div>
    <div data-reveal>
      <div class="card">
        <div class="card-lbl">Talukas — Goa State</div>
        <div class="cgrid">
          {% for t in taluks %}
          <div class="ci"><div class="cd"></div><span class="cn">{{ t }}</span></div>
          {% endfor %}
        </div>
      </div>
    </div>
  </div>
</section>

<section style="padding:55px 0 0">
  <div style="text-align:center;padding:0 80px 40px;max-width:560px;margin:0 auto">
    <div class="stag">04 / 04 — Platform</div>
    <h2 style="font-family:'DM Serif Display',serif;font-size:clamp(28px,3.8vw,46px);color:var(--ink);letter-spacing:-1.5px;margin-top:14px;line-height:1.1">Built for <em style="color:var(--teal);font-style:italic">researchers</em></h2>
  </div>
  <div class="fg">
    <div class="fi" data-reveal><div class="fi-icon">🗺️</div><div class="fi-title">Visual Risk Maps</div><p class="fi-body">Colour maps from deep green (safe) to dark red (very high risk) for every part of Goa.</p></div>
    <div class="fi" data-reveal><div class="fi-icon">🔬</div><div class="fi-title">5 Trained Models</div><p class="fi-body">Compare results across five models — agreement means high confidence in the risk reading.</p></div>
    <div class="fi" data-reveal><div class="fi-icon">📊</div><div class="fi-title">Risk Breakdown</div><p class="fi-body">See percentages for Low, Moderate, High, and Very High risk zones per taluka.</p></div>
    <div class="fi" data-reveal><div class="fi-icon">🔐</div><div class="fi-title">Secure Access</div><p class="fi-body">Separate portals for researchers and authorities with full activity tracking.</p></div>
  </div>
</section>

<div class="cta">
  <div class="cta-title">Ready to explore<br><em>Goa's risk?</em></div>
  <p class="cta-sub">Create a free researcher account or register authority credentials to access the admin panel.</p>
  <div class="cta-cards">
    <a href="{{ url_for('register') }}" class="cc cc-r">
      <div class="cc-icon">👤</div><div class="cc-title">Researcher</div>
      <div class="cc-desc">Students &amp; academics</div>
      <div class="cc-btn cb-r">Sign Up Free</div>
    </a>
    <a href="{{ url_for('login') }}" class="cc cc-r">
      <div class="cc-icon">🔑</div><div class="cc-title">Researcher Login</div>
      <div class="cc-desc">Have an account?</div>
      <div class="cc-btn cb-r">Sign In</div>
    </a>
    <a href="{{ url_for('authority_register') }}" class="cc cc-a">
      <div class="cc-icon">🏛️</div><div class="cc-title">Authority</div>
      <div class="cc-desc">Government bodies</div>
      <div class="cc-btn cb-a">Register</div>
    </a>
    <a href="{{ url_for('authority_login') }}" class="cc cc-a">
      <div class="cc-icon">🔐</div><div class="cc-title">Authority Login</div>
      <div class="cc-desc">Existing authority</div>
      <div class="cc-btn cb-a">Sign In</div>
    </a>
  </div>
</div>

<footer>
  <span>© 2024 — ML-Based Environmental Vulnerability Assessment for Goa</span>
  <span>Final Year Research Project</span>
</footer>

<script>
const nav=document.getElementById('nav');
window.addEventListener('scroll',()=>nav.classList.toggle('scrolled',window.scrollY>50));
const ro=new IntersectionObserver(e=>{e.forEach(x=>{if(x.isIntersecting){x.target.classList.add('vis');x.target.querySelectorAll('.ff').forEach(b=>setTimeout(()=>b.style.width=b.dataset.w+'%',180));}});},{threshold:0.14});
document.querySelectorAll('[data-reveal],.fi,.sc').forEach(el=>ro.observe(el));
const so=new IntersectionObserver(e=>{e.forEach(x=>{if(x.isIntersecting){x.target.classList.add('vis');const d=+(x.target.dataset.delay||0);setTimeout(()=>{const el=x.target.querySelector('[id^=cnt]');if(!el)return;const t={cnt1:12,cnt2:5,cnt3:60}[el.id];let s=null;(function r(ts){if(!s)s=ts;const p=Math.min((ts-s)/1300,1);el.textContent=Math.round(p*t);if(p<1)requestAnimationFrame(r);})(performance.now());},d);so.unobserve(x.target);}});},{threshold:0.3});
document.querySelectorAll('.sc').forEach(el=>so.observe(el));
</script>
</body>
</html>"""



# =============================================================================
#  SHARED AUTH CSS
# =============================================================================
AUTH_CSS = """
@import url('https://fonts.googleapis.com/css2?family=DM+Serif+Display:ital@0;1&family=DM+Sans:wght@300;400;500;600&family=Space+Mono:wght@400;700&display=swap');
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{
  --teal:#0d9488;--teal2:#14b8a6;--emerald:#059669;
  --ink:#0a1410;--mid:#3d5450;--soft:#7a8f8c;
  --paper:#f7faf9;--border:rgba(13,148,136,0.15);
  --admin:#1e3a5f;--admin2:#2563eb;
}
html,body{height:100%;font-family:'DM Sans',sans-serif;color:var(--ink);-webkit-font-smoothing:antialiased}
body{display:flex;align-items:center;justify-content:center;min-height:100vh;padding:20px;}
.auth-bg-user{background:linear-gradient(135deg,#edfaf7,#f7faf9,#eaf8f5);}
.auth-bg-admin{background:linear-gradient(135deg,#eef2ff,#f0f4ff,#e8f0fe);}
.auth-dots{position:fixed;inset:0;z-index:0;pointer-events:none;}
.auth-dots-user{background-image:radial-gradient(circle,rgba(13,148,136,0.06) 1px,transparent 1px);background-size:40px 40px;}
.auth-dots-admin{background-image:radial-gradient(circle,rgba(37,99,235,0.06) 1px,transparent 1px);background-size:40px 40px;}
.auth-container{position:relative;z-index:1;width:100%;max-width:460px;}
.auth-logo{text-align:center;margin-bottom:28px;animation:authUp .6s ease both;}
.auth-logo-img{width:56px;height:56px;border-radius:16px;object-fit:contain;margin-bottom:12px;box-shadow:0 10px 28px rgba(13,148,136,0.15);display:block;margin-left:auto;margin-right:auto;}
.auth-logo-icon{width:56px;height:56px;border-radius:16px;display:inline-flex;align-items:center;justify-content:center;font-size:26px;margin-bottom:12px;display:block;margin-left:auto;margin-right:auto;text-align:center;line-height:56px;}
.auth-logo-icon-user{background:linear-gradient(135deg,var(--teal),var(--emerald));box-shadow:0 10px 28px rgba(13,148,136,0.2);}
.auth-logo-icon-admin{background:linear-gradient(135deg,var(--admin),var(--admin2));box-shadow:0 10px 28px rgba(37,99,235,0.2);}
.auth-logo h1{font-family:'DM Serif Display',serif;font-size:18px;color:var(--ink);letter-spacing:-0.4px;margin-bottom:4px;line-height:1.3;}
.auth-logo p{font-family:'Space Mono',monospace;font-size:9px;letter-spacing:2px;color:var(--soft);text-transform:uppercase;}
.auth-role-tag{display:inline-flex;align-items:center;gap:6px;padding:5px 14px;border-radius:50px;font-family:'Space Mono',monospace;font-size:9px;letter-spacing:1.5px;text-transform:uppercase;margin-top:8px;}
.auth-role-tag-user{background:rgba(13,148,136,0.1);border:1px solid rgba(13,148,136,0.2);color:var(--teal);}
.auth-role-tag-admin{background:rgba(37,99,235,0.1);border:1px solid rgba(37,99,235,0.2);color:var(--admin2);}
.auth-card{background:rgba(255,255,255,0.88);border:1.5px solid rgba(255,255,255,0.6);border-radius:24px;padding:44px 40px;box-shadow:0 20px 60px rgba(10,20,16,0.1),inset 0 1px 0 rgba(255,255,255,0.9);backdrop-filter:blur(20px);animation:authUp .7s ease .1s both;position:relative;overflow:hidden;}
.auth-card::before{content:'';position:absolute;top:0;left:0;right:0;height:3px;}
.auth-card-user::before{background:linear-gradient(90deg,var(--teal),var(--emerald));}
.auth-card-admin::before{background:linear-gradient(90deg,var(--admin),var(--admin2));}
.auth-card h2{font-family:'DM Serif Display',serif;font-size:26px;color:var(--ink);margin-bottom:5px;letter-spacing:-0.5px;}
.auth-card .subtitle{font-family:'Space Mono',monospace;font-size:9px;letter-spacing:1.5px;color:var(--soft);margin-bottom:28px;text-transform:uppercase;}
.field-group{margin-bottom:18px;}
.field-group label{display:block;font-family:'Space Mono',monospace;font-size:9px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase;color:var(--mid);margin-bottom:7px;}
.field-group input,.field-group select{width:100%;padding:13px 16px;background:rgba(255,255,255,0.6);border:1.5px solid rgba(13,148,136,0.12);border-radius:13px;color:var(--ink);font-family:'DM Sans',sans-serif;font-size:15px;transition:all .3s;outline:none;}
.field-group input::placeholder{color:var(--soft);}
.field-group input:focus{background:rgba(255,255,255,0.9);box-shadow:0 0 0 4px rgba(13,148,136,0.1);}
.field-group input:focus,.field-group select:focus{border-color:var(--teal);}
.admin-field input:focus{box-shadow:0 0 0 4px rgba(37,99,235,0.1);border-color:var(--admin2);}
.submit-btn{width:100%;padding:14px;color:#fff;border:none;border-radius:13px;font-family:'DM Sans',sans-serif;font-size:15px;font-weight:600;cursor:pointer;transition:all .3s;margin-top:6px;}
.submit-btn-user{background:linear-gradient(135deg,var(--teal),var(--emerald));box-shadow:0 8px 24px rgba(13,148,136,0.25);}
.submit-btn-user:hover{transform:translateY(-2px);box-shadow:0 12px 32px rgba(13,148,136,0.35);}
.submit-btn-admin{background:linear-gradient(135deg,var(--admin),var(--admin2));box-shadow:0 8px 24px rgba(37,99,235,0.25);}
.submit-btn-admin:hover{transform:translateY(-2px);box-shadow:0 12px 32px rgba(37,99,235,0.35);}
.auth-footer{text-align:center;margin-top:20px;font-size:14px;color:var(--soft);}
.auth-footer a{color:var(--teal);text-decoration:none;font-weight:600;}
.auth-footer a.admin-link{color:var(--admin2);}
.flash-msg{padding:13px 16px;border-radius:13px;font-size:13px;margin-bottom:18px;border:1.5px solid;}
.flash-error{background:rgba(220,38,38,.05);border-color:rgba(220,38,38,.2);color:#7f1d1d;}
.flash-success{background:rgba(34,197,94,.05);border-color:rgba(34,197,94,.2);color:#15803d;}
.back-link{display:block;text-align:center;margin-top:16px;font-size:13px;color:var(--soft);text-decoration:none;transition:color .2s;}
.back-link:hover{color:var(--teal);}
.divider{display:flex;align-items:center;gap:12px;margin:18px 0;font-family:'Space Mono',monospace;font-size:9px;letter-spacing:1.5px;color:var(--soft);text-transform:uppercase;}
.divider::before,.divider::after{content:'';flex:1;height:1px;background:var(--border);}
.switch-role{display:flex;align-items:center;justify-content:center;gap:10px;padding:11px 18px;border:1.5px solid var(--border);border-radius:12px;text-decoration:none;font-size:13px;color:var(--mid);transition:all 0.25s;margin-top:4px;}
.switch-role:hover{border-color:var(--teal);color:var(--teal);}
.switch-role-admin:hover{border-color:var(--admin2);color:var(--admin2);}
.authority-badge{display:inline-flex;align-items:center;gap:6px;padding:4px 12px;background:rgba(37,99,235,0.08);border:1px solid rgba(37,99,235,0.15);border-radius:50px;font-family:'Space Mono',monospace;font-size:8px;letter-spacing:1.5px;color:var(--admin2);text-transform:uppercase;margin-bottom:16px;}
@keyframes authUp{from{opacity:0;transform:translateY(20px)}to{opacity:1;transform:translateY(0)}}
"""

# =============================================================================
#  USER LOGIN PAGE
# =============================================================================
LOGIN_HTML = """<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Researcher Sign In — EVA System</title><style>""" + AUTH_CSS + """</style></head>
<body class="auth-bg-user">
<div class="auth-dots auth-dots-user"></div>
<div class="auth-container">
  <div class="auth-logo">
    <img src="/static/images/logo.jpg" alt="EVA" class="auth-logo-img" onerror="this.outerHTML='<div class=&quot;auth-logo-icon auth-logo-icon-user&quot;>🌿</div>'">
    <h1>Environmental Vulnerability<br>Assessment — Goa</h1>
    <div class="auth-role-tag auth-role-tag-user">👤 Researcher Portal</div>
  </div>
  <div class="auth-card auth-card-user">
    <h2>Welcome back</h2>
    <p class="subtitle">Sign in to view risk assessments</p>
    {% if error %}<div class="flash-msg flash-error">{{ error }}</div>{% endif %}
    {% if success %}<div class="flash-msg flash-success">{{ success }}</div>{% endif %}
    <form method="POST" action="/login">
      <div class="field-group"><label>Username</label><input type="text" name="username" placeholder="Enter your username" required autofocus></div>
      <div class="field-group"><label>Password</label><input type="password" name="password" placeholder="Enter your password" required></div>
      <button type="submit" class="submit-btn submit-btn-user">Sign In →</button>
    </form>
    <div class="divider">or</div>
    <a href="{{ url_for('authority_login') }}" class="switch-role">🏛️ Sign in as Authority instead</a>
  </div>
  <div class="auth-footer">Don't have an account? <a href="/register">Create one</a></div>
  <a href="/" class="back-link">← Back to home</a>
</div></body></html>"""

# =============================================================================
#  USER REGISTER PAGE
# =============================================================================
REGISTER_HTML = """<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Create Account — EVA System</title><style>""" + AUTH_CSS + """</style></head>
<body class="auth-bg-user">
<div class="auth-dots auth-dots-user"></div>
<div class="auth-container">
  <div class="auth-logo">
    <img src="/static/images/logo.jpg" alt="EVA" class="auth-logo-img" onerror="this.outerHTML='<div class=&quot;auth-logo-icon auth-logo-icon-user&quot;>🌿</div>'">
    <h1>Environmental Vulnerability<br>Assessment — Goa</h1>
    <div class="auth-role-tag auth-role-tag-user">👤 Researcher Registration</div>
  </div>
  <div class="auth-card auth-card-user">
    <h2>Create account</h2>
    <p class="subtitle">Register to access risk maps</p>
    {% if error %}<div class="flash-msg flash-error">{{ error }}</div>{% endif %}
    <form method="POST" action="/register">
      <div class="field-group"><label>Full Name / Username</label><input type="text" name="username" placeholder="e.g. researcher_goa" required autofocus></div>
      <div class="field-group"><label>Email Address</label><input type="email" name="email" placeholder="your@email.com" required></div>
      <div class="field-group"><label>Password</label><input type="password" name="password" placeholder="Min. 6 characters" required></div>
      <button type="submit" class="submit-btn submit-btn-user">Create Account →</button>
    </form>
    <div class="divider">or</div>
    <a href="{{ url_for('authority_register') }}" class="switch-role">🏛️ Register as Authority instead</a>
  </div>
  <div class="auth-footer">Already registered? <a href="/login">Sign in</a></div>
  <a href="/" class="back-link">← Back to home</a>
</div></body></html>"""

# =============================================================================
#  AUTHORITY LOGIN PAGE
# =============================================================================
AUTHORITY_LOGIN_HTML = """<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Authority Sign In — EVA System</title><style>""" + AUTH_CSS + """</style></head>
<body class="auth-bg-admin">
<div class="auth-dots auth-dots-admin"></div>
<div class="auth-container">
  <div class="auth-logo">
    <div class="auth-logo-icon auth-logo-icon-admin" style="width:56px;height:56px;border-radius:16px;display:flex;align-items:center;justify-content:center;font-size:26px;margin:0 auto 12px;">🏛️</div>
    <h1>Environmental Vulnerability<br>Assessment — Goa</h1>
    <div class="auth-role-tag auth-role-tag-admin">🏛️ Authority / Admin Portal</div>
  </div>
  <div class="auth-card auth-card-admin">
    <div class="authority-badge">🔐 Restricted Access — Authority Only</div>
    <h2>Authority Sign In</h2>
    <p class="subtitle">For government bodies & administrators</p>
    {% if error %}<div class="flash-msg flash-error">{{ error }}</div>{% endif %}
    {% if success %}<div class="flash-msg flash-success">{{ success }}</div>{% endif %}
    <form method="POST" action="/authority/login">
      <div class="field-group admin-field"><label>Authority Username</label><input type="text" name="username" placeholder="Enter authority username" required autofocus></div>
      <div class="field-group admin-field"><label>Password</label><input type="password" name="password" placeholder="Enter password" required></div>
      <button type="submit" class="submit-btn submit-btn-admin">Authority Sign In →</button>
    </form>
    <div class="divider">or</div>
    <a href="{{ url_for('login') }}" class="switch-role switch-role-admin">👤 Sign in as Researcher instead</a>
  </div>
  <div class="auth-footer">Need authority access? <a href="{{ url_for('authority_register') }}" class="admin-link">Register here</a></div>
  <a href="/" class="back-link">← Back to home</a>
</div></body></html>"""

# =============================================================================
#  AUTHORITY REGISTER PAGE
# =============================================================================
AUTHORITY_REGISTER_HTML = """<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Authority Registration — EVA System</title><style>""" + AUTH_CSS + """</style></head>
<body class="auth-bg-admin">
<div class="auth-dots auth-dots-admin"></div>
<div class="auth-container" style="max-width:500px">
  <div class="auth-logo">
    <div class="auth-logo-icon auth-logo-icon-admin" style="width:56px;height:56px;border-radius:16px;display:flex;align-items:center;justify-content:center;font-size:26px;margin:0 auto 12px;">🏛️</div>
    <h1>Environmental Vulnerability<br>Assessment — Goa</h1>
    <div class="auth-role-tag auth-role-tag-admin">🏛️ Authority Registration</div>
  </div>
  <div class="auth-card auth-card-admin">
    <div class="authority-badge">🔐 Authority / Government Access</div>
    <h2>Authority Registration</h2>
    <p class="subtitle">For official bodies & administrators only</p>
    {% if error %}<div class="flash-msg flash-error">{{ error }}</div>{% endif %}
    <form method="POST" action="/authority/register">
      <div class="field-group admin-field"><label>Full Name</label><input type="text" name="username" placeholder="e.g. collector_north_goa" required autofocus></div>
      <div class="field-group admin-field"><label>Official Email</label><input type="email" name="email" placeholder="official@gov.in" required></div>
      <div class="field-group admin-field"><label>Department / Organisation</label><input type="text" name="department" placeholder="e.g. Goa State Disaster Management Authority"></div>
      <div class="field-group admin-field"><label>Password</label><input type="password" name="password" placeholder="Min. 6 characters" required></div>
      <div class="field-group admin-field"><label>Authority Access Code</label><input type="password" name="access_code" placeholder="Enter the authority access code" required></div>
      <button type="submit" class="submit-btn submit-btn-admin">Register as Authority →</button>
    </form>
    <div class="divider">or</div>
    <a href="{{ url_for('register') }}" class="switch-role switch-role-admin">👤 Register as Researcher instead</a>
  </div>
  <div class="auth-footer">Already registered? <a href="{{ url_for('authority_login') }}" class="admin-link">Authority Sign In</a></div>
  <a href="/" class="back-link">← Back to home</a>
</div></body></html>"""

# =============================================================================
#  DASHBOARD CSS
# =============================================================================
DASHBOARD_CSS = """
@import url('https://fonts.googleapis.com/css2?family=DM+Serif+Display:ital@0;1&family=DM+Sans:wght@300;400;500;600;700&family=Space+Mono:wght@400;700&display=swap');
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{
  --teal:#0d9488;--teal2:#14b8a6;--emerald:#059669;
  --ink:#0a1410;--mid:#3d5450;--soft:#7a8f8c;
  --paper:#f7faf9;--border:rgba(13,148,136,0.15);
  --glass:rgba(247,250,249,0.85);--sidebar-w:272px;
  --admin:#1e3a5f;--admin2:#2563eb;
}
html{scroll-behavior:smooth}
body{font-family:'DM Sans',sans-serif;background:linear-gradient(135deg,#f0faf8,#f7faf9);color:var(--ink);min-height:100vh;-webkit-font-smoothing:antialiased;display:flex}
.sidebar{width:var(--sidebar-w);min-height:100vh;background:rgba(255,255,255,0.9);border-right:1.5px solid var(--border);display:flex;flex-direction:column;padding:28px 0;position:fixed;top:0;left:0;bottom:0;z-index:50;box-shadow:4px 0 28px rgba(13,148,136,0.07);backdrop-filter:blur(20px);}
.sidebar-logo{display:flex;align-items:center;gap:12px;padding:0 20px 24px;border-bottom:1.5px solid var(--border);margin-bottom:22px;}
.sidebar-logo-img{width:38px;height:38px;border-radius:11px;object-fit:contain;box-shadow:0 5px 14px rgba(13,148,136,0.15);flex-shrink:0;}
.sidebar-logo-icon{width:38px;height:38px;background:linear-gradient(135deg,var(--teal),var(--emerald));border-radius:11px;display:flex;align-items:center;justify-content:center;font-size:17px;flex-shrink:0;}
.sidebar-brand{font-family:'DM Serif Display',serif;font-size:13px;color:var(--ink);letter-spacing:-0.2px;line-height:1.35;}
.sidebar-brand span{display:block;font-family:'Space Mono',monospace;font-size:8px;color:var(--soft);font-weight:400;margin-top:3px;letter-spacing:1.5px;text-transform:uppercase;}
.nav-section-label{font-family:'Space Mono',monospace;font-size:8px;letter-spacing:2px;text-transform:uppercase;color:var(--soft);padding:0 20px;margin-bottom:6px;font-weight:700;}
.nav-item{display:flex;align-items:center;gap:10px;padding:10px 20px;color:var(--mid);text-decoration:none;font-size:14px;font-weight:500;transition:all .2s;border-left:2px solid transparent;margin:2px 0;}
.nav-item:hover{color:var(--teal);background:rgba(13,148,136,.06);transform:translateX(2px);}
.nav-item.active{color:var(--teal);background:linear-gradient(90deg,rgba(13,148,136,.1),transparent);border-left-color:var(--teal);font-weight:600;}
.nav-item .icon{width:17px;text-align:center;font-size:14px;flex-shrink:0;}
.sidebar-spacer{flex:1;}
.sidebar-user{padding:16px 20px 0;border-top:1.5px solid var(--border);}
.sidebar-user-info{display:flex;align-items:center;gap:10px;margin-bottom:10px;}
.sidebar-avatar{width:34px;height:34px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:13px;font-weight:700;color:white;flex-shrink:0;}
.avatar-user{background:linear-gradient(135deg,var(--teal),var(--emerald));}
.avatar-admin{background:linear-gradient(135deg,var(--admin),var(--admin2));}
.sidebar-user-name{font-size:13px;font-weight:600;color:var(--ink);}
.sidebar-user-role{font-family:'Space Mono',monospace;font-size:8px;padding:2px 7px;border-radius:50px;display:inline-block;margin-top:2px;letter-spacing:1px;}
.role-user{color:var(--teal);background:rgba(13,148,136,.1);border:1px solid rgba(13,148,136,.2);}
.role-admin{color:var(--admin2);background:rgba(37,99,235,.08);border:1px solid rgba(37,99,235,.15);}
.logout-btn{display:flex;align-items:center;gap:7px;width:100%;padding:9px 11px;background:rgba(220,38,38,.05);border:1.5px solid rgba(220,38,38,.1);border-radius:10px;color:#b91c1c;font-size:13px;font-weight:600;cursor:pointer;text-decoration:none;transition:all .2s;}
.logout-btn:hover{background:rgba(220,38,38,.09);}
.main{margin-left:var(--sidebar-w);flex:1;min-height:100vh;display:flex;flex-direction:column;}
.topbar{padding:18px 36px;border-bottom:1.5px solid var(--border);display:flex;align-items:center;justify-content:space-between;background:rgba(255,255,255,0.9);backdrop-filter:blur(20px);position:sticky;top:0;z-index:40;box-shadow:0 4px 18px rgba(13,148,136,.05);}
.topbar-title{font-family:'DM Serif Display',serif;font-size:18px;color:var(--ink);letter-spacing:-0.4px;}
.topbar-title span{font-family:'Space Mono',monospace;font-size:9px;color:var(--soft);font-weight:400;display:block;margin-top:2px;letter-spacing:1.5px;text-transform:uppercase;}
.topbar-badge{padding:5px 13px;border-radius:50px;font-family:'Space Mono',monospace;font-size:8px;font-weight:700;border:1px solid;letter-spacing:1px;text-transform:uppercase;}
.badge-admin{background:rgba(37,99,235,.07);border-color:rgba(37,99,235,.15);color:var(--admin2);}
.badge-user{background:rgba(13,148,136,.1);border-color:rgba(13,148,136,.2);color:var(--teal);}
.content{padding:36px;flex:1;animation:dashIn .5s ease both;}
@keyframes dashIn{from{opacity:0;transform:translateY(12px)}to{opacity:1;transform:translateY(0)}}
.dash-hero{position:relative;border-radius:26px;overflow:hidden;margin-bottom:32px;min-height:240px;display:flex;align-items:flex-end;padding:36px;background:linear-gradient(135deg,#0a1f18,#0f3d2a,#1a5c3a);box-shadow:0 18px 50px rgba(13,148,136,.15);}
.dash-hero-bg{position:absolute;inset:0;background-image:url('/static/images/goa_map_bg.jpg');background-size:cover;background-position:center;opacity:0.5;z-index:0;}
.dash-hero-overlay{position:absolute;inset:0;background:linear-gradient(to bottom,rgba(10,25,18,.1) 0%,rgba(10,25,18,.85) 100%);z-index:1;}
.dash-hero-content{position:relative;z-index:2;}
.dash-hero h2{font-family:'DM Serif Display',serif;font-size:30px;color:#fff;letter-spacing:-0.7px;margin-bottom:8px;text-shadow:0 3px 18px rgba(0,0,0,.5);}
.dash-hero p{font-size:14px;color:rgba(255,255,255,.82);max-width:500px;line-height:1.75;}
.dash-hero-chips{display:flex;gap:10px;margin-top:18px;flex-wrap:wrap;}
.dash-hero-chip{padding:6px 16px;border-radius:50px;background:rgba(255,255,255,.13);border:1px solid rgba(255,255,255,.22);color:rgba(255,255,255,.9);font-family:'Space Mono',monospace;font-size:8px;letter-spacing:1.5px;backdrop-filter:blur(10px);}
.form-card{background:rgba(255,255,255,.9);border:1.5px solid rgba(255,255,255,.6);border-radius:22px;padding:36px;max-width:660px;box-shadow:0 14px 44px rgba(13,148,136,.09),inset 0 1px 0 rgba(255,255,255,.9);backdrop-filter:blur(20px);position:relative;overflow:hidden;}
.form-card::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;background:linear-gradient(90deg,var(--teal),var(--emerald),var(--teal2));background-size:200% auto;animation:shimmer 3s linear infinite;}
@keyframes shimmer{0%{background-position:-200% center}100%{background-position:200% center}}
.form-card-header{display:flex;align-items:center;gap:13px;margin-bottom:28px;}
.form-card-icon{width:46px;height:46px;background:rgba(13,148,136,.1);border:1.5px solid rgba(13,148,136,.2);border-radius:13px;display:flex;align-items:center;justify-content:center;font-size:20px;}
.form-card-title{font-family:'DM Serif Display',serif;font-size:19px;color:var(--ink);letter-spacing:-0.4px;}
.form-card-subtitle{font-family:'Space Mono',monospace;font-size:9px;color:var(--soft);margin-top:3px;letter-spacing:1px;}
.form-row{display:grid;grid-template-columns:1fr 1fr;gap:18px;margin-bottom:22px;}
.field-group label{display:block;font-family:'Space Mono',monospace;font-size:9px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase;color:var(--mid);margin-bottom:7px;}
.field-group select{width:100%;padding:12px 15px;background:rgba(255,255,255,.6);border:1.5px solid rgba(13,148,136,.12);border-radius:12px;color:var(--ink);font-family:'DM Sans',sans-serif;font-size:14px;transition:all .3s;outline:none;appearance:none;cursor:pointer;background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' fill='none' viewBox='0 0 20 20'%3E%3Cpath stroke='%236b7a78' stroke-width='1.5' d='M6 8l4 4 4-4'/%3E%3C/svg%3E");background-repeat:no-repeat;background-position:right 12px center;background-size:16px;padding-right:36px;}
.field-group select:focus{border-color:var(--teal);box-shadow:0 0 0 4px rgba(13,148,136,.1);}
.run-btn{width:100%;padding:13px;background:linear-gradient(135deg,var(--teal),var(--emerald));color:#fff;border:none;border-radius:12px;font-family:'DM Sans',sans-serif;font-size:15px;font-weight:600;cursor:pointer;transition:all .3s;box-shadow:0 7px 22px rgba(13,148,136,.24);}
.run-btn:hover{transform:translateY(-2px);box-shadow:0 12px 30px rgba(13,148,136,.35);}
.spinner-overlay{display:none;position:fixed;inset:0;z-index:200;background:rgba(247,250,249,.92);backdrop-filter:blur(8px);align-items:center;justify-content:center;flex-direction:column;gap:20px;}
.spinner-overlay.active{display:flex;}
.spinner{width:50px;height:50px;border:2px solid rgba(13,148,136,.15);border-top-color:var(--teal);border-radius:50%;animation:spin .8s linear infinite;}
@keyframes spin{to{transform:rotate(360deg)}}
.spinner-text{font-family:'Space Mono',monospace;font-size:10px;letter-spacing:2px;color:var(--soft);text-transform:uppercase;}
.result-card{background:rgba(255,255,255,.9);border:1.5px solid rgba(255,255,255,.6);border-radius:22px;padding:30px;max-width:900px;margin-top:28px;animation:dashIn .5s ease both;box-shadow:0 14px 44px rgba(13,148,136,.09);backdrop-filter:blur(20px);}
.result-header{display:flex;align-items:flex-start;justify-content:space-between;margin-bottom:22px;flex-wrap:wrap;gap:10px;}
.result-title{font-family:'DM Serif Display',serif;font-size:20px;color:var(--ink);letter-spacing:-0.4px;}
.result-meta{font-family:'Space Mono',monospace;font-size:9px;color:var(--soft);margin-top:3px;letter-spacing:1px;}
.badge-row{display:flex;gap:8px;flex-wrap:wrap;}
.badge{padding:5px 12px;border-radius:50px;font-family:'Space Mono',monospace;font-size:8px;font-weight:700;border:1px solid;letter-spacing:1px;text-transform:uppercase;}
.badge-teal{background:rgba(13,148,136,.1);border-color:rgba(13,148,136,.2);color:var(--teal);}
.badge-blue{background:rgba(59,130,246,.07);border-color:rgba(59,130,246,.15);color:#1d4ed8;}
.badge-em{background:rgba(34,197,94,.07);border-color:rgba(34,197,94,.15);color:#15803d;}
.result-img{width:100%;border-radius:14px;border:1.5px solid var(--border);display:block;box-shadow:0 7px 28px rgba(13,148,136,.09);}
.result-footer{display:flex;align-items:center;justify-content:space-between;margin-top:18px;flex-wrap:wrap;gap:10px;}
.file-name{font-family:'Space Mono',monospace;font-size:10px;color:var(--soft);letter-spacing:1px;}
.back-link-btn{display:inline-flex;align-items:center;gap:6px;padding:8px 16px;background:rgba(13,148,136,.1);border:1.5px solid rgba(13,148,136,.2);border-radius:10px;color:var(--teal);font-size:13px;font-weight:600;text-decoration:none;transition:all .2s;}
.back-link-btn:hover{background:rgba(13,148,136,.16);transform:translateY(-1px);}
.stats-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(190px,1fr));gap:18px;margin-bottom:28px;}
.stat-card{background:rgba(255,255,255,.9);border:1.5px solid rgba(255,255,255,.6);border-radius:16px;padding:26px;transition:all .3s;box-shadow:0 4px 14px rgba(13,148,136,.06);backdrop-filter:blur(10px);}
.stat-card:hover{box-shadow:0 10px 36px rgba(13,148,136,.13);transform:translateY(-3px);}
.stat-card-label{font-family:'Space Mono',monospace;font-size:9px;letter-spacing:1.5px;text-transform:uppercase;color:var(--soft);margin-bottom:10px;}
.stat-card-value{font-family:'DM Serif Display',serif;font-size:42px;color:var(--teal);letter-spacing:-1px;}
.table-wrap{background:rgba(255,255,255,.9);border:1.5px solid rgba(255,255,255,.6);border-radius:16px;overflow:hidden;box-shadow:0 4px 14px rgba(13,148,136,.06);backdrop-filter:blur(10px);}
.table-head{padding:15px 20px;border-bottom:1.5px solid var(--border);font-family:'Space Mono',monospace;font-size:10px;letter-spacing:1.5px;text-transform:uppercase;color:var(--mid);}
table{width:100%;border-collapse:collapse;}
th{padding:11px 18px;text-align:left;font-family:'Space Mono',monospace;font-size:8px;letter-spacing:1.5px;text-transform:uppercase;color:var(--soft);font-weight:700;}
td{padding:12px 18px;font-size:13px;color:var(--mid);border-top:1px solid var(--border);}
tr:hover td{background:rgba(13,148,136,.02);}
.error-card{background:rgba(220,38,38,.05);border:1.5px solid rgba(220,38,38,.15);border-radius:14px;padding:26px 30px;max-width:560px;}
.error-card h3{font-family:'DM Serif Display',serif;color:#b91c1c;margin-bottom:10px;font-size:20px;}
.error-card p{font-size:14px;color:#7f1d1d;line-height:1.7;}
code{display:inline-block;background:rgba(220,38,38,.08);padding:4px 9px;border-radius:6px;font-family:'Space Mono',monospace;font-size:10px;word-break:break-all;color:#991b1b;margin-top:8px;letter-spacing:0.5px;}
.model-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:12px;margin-bottom:26px;}
.model-select-card{padding:16px;border-radius:14px;border:1.5px solid var(--border);background:rgba(255,255,255,.6);cursor:pointer;transition:all .3s;text-align:center;}
.model-select-card:hover{border-color:rgba(13,148,136,.3);background:rgba(13,148,136,.05);transform:translateY(-3px);box-shadow:0 7px 22px rgba(13,148,136,.1);}
.model-select-card.selected{border-color:var(--teal);background:rgba(13,148,136,.1);box-shadow:0 0 0 3px rgba(13,148,136,.08);}
.model-select-card .m-icon{font-size:22px;margin-bottom:7px;}
.model-select-card .m-name{font-family:'DM Serif Display',serif;font-size:17px;color:var(--ink);}
.model-select-card .m-full{font-family:'Space Mono',monospace;font-size:8px;color:var(--soft);margin-top:3px;letter-spacing:1px;}
"""

# =============================================================================
#  DASHBOARD LAYOUT
# =============================================================================
LAYOUT_START = """<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>{{ page_title }} — EVA System</title>
<style>""" + DASHBOARD_CSS + """</style></head>
<body>
<aside class="sidebar">
  <div class="sidebar-logo">
    <img src="/static/images/logo.jpg" alt="EVA" class="sidebar-logo-img" onerror="this.outerHTML='<div class=sidebar-logo-icon>🌿</div>'">
    <div>
      <div class="sidebar-brand">ML-Based Environmental<br>Vulnerability Assessment<span>Goa · Landslide Risk</span></div>
    </div>
  </div>
  <div class="nav-section-label">Navigation</div>
  <a href="/dashboard" class="nav-item {% if active=='assess' %}active{% endif %}"><span class="icon">🗺️</span> Area Risk Assessment</a>
  <a href="/goa-overview" class="nav-item {% if active=='goa' %}active{% endif %}"><span class="icon">🌏</span> Full Goa Risk Map</a>
  {% if role=='admin' %}
  <div class="nav-section-label" style="margin-top:14px">Authority Panel</div>
  <a href="/admin" class="nav-item {% if active=='admin' %}active{% endif %}"><span class="icon">📊</span> Dashboard</a>
  <a href="/admin/logs" class="nav-item {% if active=='logs' %}active{% endif %}"><span class="icon">📋</span> Activity Logs</a>
  <a href="/admin/users" class="nav-item {% if active=='users' %}active{% endif %}"><span class="icon">👥</span> Users</a>
  {% endif %}
  <div class="sidebar-spacer"></div>
  <div class="sidebar-user">
    <div class="sidebar-user-info">
      <div class="sidebar-avatar {% if role=='admin' %}avatar-admin{% else %}avatar-user{% endif %}">{{ username[0].upper() }}</div>
      <div>
        <div class="sidebar-user-name">{{ username }}</div>
        <div class="sidebar-user-role {% if role=='admin' %}role-admin{% else %}role-user{% endif %}">{% if role=='admin' %}🏛️ Authority{% else %}👤 Researcher{% endif %}</div>
      </div>
    </div>
    <a href="/logout" class="logout-btn">⬡ Sign Out</a>
  </div>
</aside>
<main class="main">
  <div class="topbar">
    <div class="topbar-title">{{ page_title }}<span>{{ page_subtitle }}</span></div>
    <div class="topbar-badge {% if role=='admin' %}badge-admin{% else %}badge-user{% endif %}">
      {% if role=='admin' %}🏛️ Authority{% else %}👤 Researcher{% endif %}
    </div>
  </div>
  <div class="content">"""

LAYOUT_END = """
  </div>
</main>
<div class="spinner-overlay" id="spinner">
  <div class="spinner"></div>
  <div class="spinner-text">Generating Risk Map</div>
</div>
<script>
document.addEventListener('DOMContentLoaded',function(){
  const f=document.getElementById('assessForm');
  if(f) f.addEventListener('submit',function(){document.getElementById('spinner').classList.add('active');});
  document.querySelectorAll('.model-select-card').forEach(card=>{
    card.addEventListener('click',function(){
      document.querySelectorAll('.model-select-card').forEach(c=>c.classList.remove('selected'));
      this.classList.add('selected');
      const h=document.getElementById('model_hidden');
      if(h) h.value=this.dataset.model;
    });
  });
});
</script>
</body></html>"""

def render_layout(body, page_title, page_subtitle, active):
    return render_template_string(
        LAYOUT_START + body + LAYOUT_END,
        page_title=page_title, page_subtitle=page_subtitle, active=active,
        username=session['username'], role=session['role']
    )

# =============================================================================
#  ROUTES
# =============================================================================
AUTHORITY_ACCESS_CODE = "GOA2024AUTH"  # Change this to your actual authority access code

@app.route('/')
def home():
    return render_template_string(LANDING_HTML, taluks=taluks,
                                   model_names=MODEL_FULL_NAMES, icons=MODEL_ICONS)

# ── USER AUTH ──────────────────────────────────────────────────────────────────
@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard') if session.get('role') != 'admin' else url_for('admin_dashboard'))
    error = None
    success = "Account created. Please sign in." if request.args.get('registered') else None
    if request.method == 'POST':
        username = request.form.get('username','').strip()
        password = request.form.get('password','')
        conn = get_db()
        try:
            user = conn.execute("SELECT * FROM users WHERE username=? AND role='user'", (username,)).fetchone()
        finally:
            conn.close()
        if user and user['password_hash'] == hash_password(password):
            session['user_id']  = user['id']
            session['username'] = user['username']
            session['role']     = user['role']
            return redirect(url_for('dashboard'))
        error = "Invalid username or password."
    return render_template_string(LOGIN_HTML, error=error, success=success)

@app.route('/register', methods=['GET', 'POST'])
def register():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    error = None
    if request.method == 'POST':
        username = request.form.get('username','').strip()
        email    = request.form.get('email','').strip()
        password = request.form.get('password','')
        if len(password) < 6:
            error = "Password must be at least 6 characters."
        elif not username:
            error = "Username is required."
        else:
            conn = get_db(); now = datetime.now().isoformat()
            try:
                conn.execute("INSERT INTO users (username,email,password_hash,role,created_at) VALUES (?,?,?,?,?)",
                             (username, email, hash_password(password), 'user', now))
                conn.commit()
                append_user_to_excel(username, email, 'user', now[:16])
                return redirect(url_for('login') + '?registered=1')
            except sqlite3.IntegrityError:
                error = "Username or email already exists."
            finally:
                conn.close()
    return render_template_string(REGISTER_HTML, error=error)

# ── AUTHORITY AUTH ─────────────────────────────────────────────────────────────
@app.route('/authority/login', methods=['GET', 'POST'])
def authority_login():
    if 'user_id' in session:
        return redirect(url_for('admin_dashboard') if session.get('role') == 'admin' else url_for('dashboard'))
    error = None
    success = "Authority account created. Please sign in." if request.args.get('registered') else None
    if request.method == 'POST':
        username = request.form.get('username','').strip()
        password = request.form.get('password','')
        conn = get_db()
        try:
            user = conn.execute("SELECT * FROM users WHERE username=? AND role='admin'", (username,)).fetchone()
        finally:
            conn.close()
        if user and user['password_hash'] == hash_password(password):
            session['user_id']  = user['id']
            session['username'] = user['username']
            session['role']     = user['role']
            return redirect(url_for('admin_dashboard'))
        error = "Invalid authority credentials."
    return render_template_string(AUTHORITY_LOGIN_HTML, error=error, success=success)

@app.route('/authority/register', methods=['GET', 'POST'])
def authority_register():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    error = None
    if request.method == 'POST':
        username    = request.form.get('username','').strip()
        email       = request.form.get('email','').strip()
        password    = request.form.get('password','')
        access_code = request.form.get('access_code','').strip()
        department  = request.form.get('department','').strip()
        if access_code != AUTHORITY_ACCESS_CODE:
            error = "Invalid authority access code. Please contact the system administrator."
        elif len(password) < 6:
            error = "Password must be at least 6 characters."
        elif not username:
            error = "Username is required."
        else:
            conn = get_db(); now = datetime.now().isoformat()
            try:
                conn.execute("INSERT INTO users (username,email,password_hash,role,created_at) VALUES (?,?,?,?,?)",
                             (username, email, hash_password(password), 'admin', now))
                conn.commit()
                append_user_to_excel(username, email, 'admin (authority)', now[:16])
                return redirect(url_for('authority_login') + '?registered=1')
            except sqlite3.IntegrityError:
                error = "Username or email already exists."
            finally:
                conn.close()
    return render_template_string(AUTHORITY_REGISTER_HTML, error=error)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('home'))


@app.route('/dashboard')
@login_required
def dashboard():
    taluk_opts = "".join(f'<option value="{t}">{t}</option>' for t in taluks)
    model_opts = "".join(f'<option value="{m}">{MODEL_FULL_NAMES[m]}</option>' for m in models)
    body = f"""
<div class="dash-hero">
  <div class="dash-hero-bg"></div>
  <div class="dash-hero-overlay"></div>
  <div class="dash-hero-content">
    <h2>Landslide Risk Assessment</h2>
    <p>Choose any of the 12 areas of Goa and compare results across 5 Machine Learning models to see where the high-risk zones are.</p>
    <div class="dash-hero-chips">
      <span class="dash-hero-chip">🗺️ 12 AREAS</span>
      <span class="dash-hero-chip">🔬 5 MODELS</span>
      <span class="dash-hero-chip">📊 RISK MAPS</span>
    </div>
  </div>
</div>
<div class="form-card">
  <div class="form-card-header">
    <div class="form-card-icon">🗺️</div>
    <div>
      <div class="form-card-title">Check Area Risk</div>
      <div class="form-card-subtitle">Pick a taluka and model to generate the risk map</div>
    </div>
  </div>
  <form id="assessForm" action="/show" method="POST">
    <div class="form-row">
      <div class="field-group"><label>Taluka</label><select name="taluk">{taluk_opts}</select></div>
      <div class="field-group"><label>Model</label><select name="model">{model_opts}</select></div>
    </div>
    <button type="submit" class="run-btn">Generate Risk Map →</button>
  </form>
</div>"""
    return render_layout(body, "Area Risk Assessment", "Select a taluka and model to view the risk map", "assess")

@app.route('/goa-overview', methods=['GET', 'POST'])
@login_required
def goa_overview():
    from rasterio.warp import reproject as _warp, Resampling, transform_bounds
    from rasterio.crs import CRS as _CRS
    from rasterio.transform import from_bounds as _from_bounds

    WGS84 = _CRS.from_epsg(4326)
    image_b64 = selected_model = filename = error_msg = None

    if request.method == 'POST':
        selected_model = request.form.get('model','').strip()
        if selected_model in models:
            file_path, filename = get_goa_file_path(selected_model)
            if not file_path or not os.path.exists(file_path):
                sf, fn = GOA_OVERVIEW_FILES.get(selected_model, ('?','?'))
                error_msg = (f"Risk map not found for <strong>{selected_model}</strong>.<br>"
                             f"Expected file: <code>{os.path.join(BASE_DIR,GOA_ML_FOLDER,sf,fn)}</code>")
            else:
                gdf = get_gdf()
                if gdf is not None and len(gdf) > 0:
                    sb = gdf.total_bounds
                    ml, mb_, mr, mt = sb[0], sb[1], sb[2], sb[3]
                else:
                    ml, mb_, mr, mt = None, None, None, None

                with rasterio.open(file_path) as src:
                    src_crs = src.crs; nd = src.nodata
                    oh, ow = 600, 600
                    if src_crs and src_crs.to_epsg() != 4326:
                        rbl, rbb, rbr, rbt = transform_bounds(src_crs, WGS84,
                            src.bounds.left, src.bounds.bottom, src.bounds.right, src.bounds.top)
                    else:
                        rbl, rbb, rbr, rbt = (src.bounds.left, src.bounds.bottom,
                                               src.bounds.right, src.bounds.top)
                    if ml is None:
                        ml, mb_, mr, mt = rbl, rbb, rbr, rbt
                    dst_t = _from_bounds(ml, mb_, mr, mt, ow, oh)
                    raw = np.full((oh, ow), np.nan, dtype=np.float32)
                    _warp(source=rasterio.band(src, 1), destination=raw,
                          src_transform=src.transform, src_crs=src_crs,
                          dst_transform=dst_t, dst_crs=WGS84,
                          resampling=Resampling.bilinear, src_nodata=nd, dst_nodata=np.nan)
                    raw = raw.astype(float)
                    if nd is not None: raw[raw == float(nd)] = np.nan

                vmin=float(np.nanpercentile(raw[~np.isnan(raw)],2)) if np.any(~np.isnan(raw)) else 0
                vmax=float(np.nanpercentile(raw[~np.isnan(raw)],98)) if np.any(~np.isnan(raw)) else 1

                fig,ax=plt.subplots(figsize=(8,9),facecolor='#ffffff')
                ax.set_facecolor('#f8fffe')
                img=ax.imshow(raw,extent=[ml,mr,mb_,mt],cmap=RISK_CMAP,vmin=vmin,vmax=vmax,
                              origin='upper',aspect='auto',zorder=2,alpha=0.92,interpolation='bilinear')
                if gdf is not None:
                    for _, row in gdf.iterrows():
                        geom = row.geometry
                        gs = [geom] if geom.geom_type == 'Polygon' else list(geom.geoms)
                        for g in gs:
                            xs, ys = g.exterior.xy
                            ax.plot(list(xs)+[xs[0]], list(ys)+[ys[0]], color='white', linewidth=3.0, zorder=5, alpha=0.7)
                            ax.plot(list(xs)+[xs[0]], list(ys)+[ys[0]], color='#0a4a42', linewidth=1.8, zorder=6)
                for taluka_name, (cx, cy) in TALUKA_CENTROIDS.items():
                    ax.text(cx, cy, taluka_name, ha='center', va='center',
                            fontsize=7.5, color='#0f2027', fontweight='bold',
                            bbox=dict(boxstyle='round,pad=0.28', facecolor='white',
                                      edgecolor='#0a4a42', alpha=0.92, linewidth=1.0), zorder=8)
                padx=(mr-ml)*0.05; pady=(mt-mb_)*0.05
                ax.set_xlim(ml-padx,mr+padx); ax.set_ylim(mb_-pady,mt+pady)
                cbar=plt.colorbar(img,ax=ax,fraction=0.025,pad=0.02,shrink=0.55)
                cbar.ax.tick_params(colors='#475569',labelsize=8)
                cbar.outline.set_edgecolor('#e2e8f0')
                cbar.set_label('Risk Level',color='#475569',fontsize=9)
                ax.set_title(f"Goa — Full State  |  {MODEL_FULL_NAMES.get(selected_model,selected_model)}",
                             fontsize=11,pad=14,color='#0f2027')
                ax.tick_params(colors='#94a3b8',labelsize=7)
                for sp in ax.spines.values(): sp.set_edgecolor('#e2e8f0')
                ax.set_xlabel('Longitude',color='#94a3b8',fontsize=8)
                ax.set_ylabel('Latitude',color='#94a3b8',fontsize=8)
                fig.patch.set_facecolor('#ffffff'); plt.tight_layout()
                buf=io.BytesIO()
                plt.savefig(buf,format='png',dpi=110,bbox_inches='tight',facecolor='#ffffff')
                buf.seek(0); plt.close(fig)
                image_b64=base64.b64encode(buf.getvalue()).decode()
                log_assessment(session['user_id'],session['username'],'Goa (Full State)',selected_model)

    cards="".join(f"""
    <div class="model-select-card {'selected' if m==selected_model else ''}" data-model="{m}">
      <div class="m-icon">{MODEL_ICONS.get(m,'🤖')}</div>
      <div class="m-name">{m}</div>
      <div class="m-full">{MODEL_FULL_NAMES[m].split('—')[1].strip()}</div>
    </div>""" for m in models)

    result_html=""
    if error_msg:
        result_html=f'<div class="error-card" style="margin-top:24px"><h3>⚠ File Not Found</h3><p>{error_msg}</p></div>'
    elif image_b64:
        now_str=datetime.now().strftime("%d %b %Y, %H:%M")
        result_html=f"""
<div class="result-card">
  <div class="result-header">
    <div><div class="result-title">Goa State — Full Risk Overview</div>
    <div class="result-meta">Generated {now_str} · {session['username']}</div></div>
    <div class="badge-row">
      <span class="badge badge-teal">🌏 All 12 Areas</span>
      <span class="badge badge-blue">⚙ {MODEL_FULL_NAMES.get(selected_model,'')}</span>
      <span class="badge badge-em">Risk Map</span>
    </div>
  </div>
  <img class="result-img" src="data:image/png;base64,{image_b64}" alt="Full Goa Risk Map">
  <div class="result-footer"><span class="file-name">{filename}</span><a href="/goa-overview" class="back-link-btn">← New Map</a></div>
</div>"""

    body=f"""
<div class="form-card" style="max-width:680px">
  <div class="form-card-header">
    <div class="form-card-icon">🌏</div>
    <div><div class="form-card-title">Full Goa Risk Map</div>
    <div class="form-card-subtitle">Select an model to see risk across the whole state</div></div>
  </div>
  <form id="assessForm" action="/goa-overview" method="POST">
    <input type="hidden" name="model" id="model_hidden" value="{selected_model or ''}">
    <p style="font-family:'Space Mono',monospace;font-size:9px;letter-spacing:1.5px;text-transform:uppercase;color:var(--soft);margin-bottom:12px">Choose Model</p>
    <div class="model-grid">{cards}</div>
    <button type="submit" class="run-btn">Generate Full Goa Risk Map →</button>
  </form>
</div>
{result_html}
<script>
document.querySelectorAll('.model-select-card').forEach(el=>{{
  el.addEventListener('click',function(){{
    document.querySelectorAll('.model-select-card').forEach(c=>c.classList.remove('selected'));
    this.classList.add('selected');
    document.getElementById('model_hidden').value=this.dataset.model;
  }});
}});
</script>"""
    return render_layout(body,"Full Goa Risk Map","See risk levels across all areas of Goa","goa")

@app.route('/show', methods=['POST'])
@login_required
def show():
    from rasterio.warp import reproject as _warp, Resampling, transform_bounds
    from rasterio.crs import CRS as _CRS
    from rasterio.transform import from_bounds as _from_bounds
    WGS84 = _CRS.from_epsg(4326)

    taluk=request.form.get('taluk','').strip()
    model=request.form.get('model','').strip()
    if not taluk or not model: return redirect(url_for('dashboard'))

    log_assessment(session['user_id'],session['username'],taluk,model)
    file_path,filename=get_file_path(taluk,model)

    if not os.path.exists(file_path):
        body=f"""<div class="error-card"><h3>⚠ Map Not Found</h3>
        <p>No risk map found for <strong>{taluk}</strong> using <strong>{model}</strong>.</p>
        <p>Expected file location:</p><code>{file_path}</code></div>
        <br><a href="/dashboard" class="back-link-btn">← Go Back</a>"""
        return render_layout(body,"Map Not Found","","assess")

    gdf=get_gdf(); ncol=gdf_name_col(gdf) if gdf is not None else None
    target=SHAPEFILE_NAME_MAP.get(taluk,taluk)

    selected_geom = None
    if gdf is not None and ncol:
        for _, row in gdf.iterrows():
            if row[ncol] == target:
                selected_geom = row.geometry; break

    if gdf is not None and len(gdf) > 0:
        sb = gdf.total_bounds
        canvas_l, canvas_b, canvas_r, canvas_t = sb[0], sb[1], sb[2], sb[3]
    else:
        canvas_l = canvas_b = canvas_r = canvas_t = None

    with rasterio.open(file_path) as src:
        src_crs = src.crs; nd = src.nodata
        oh, ow = 800, 800
        if canvas_l is None:
            if src_crs and src_crs.to_epsg() != 4326:
                canvas_l, canvas_b, canvas_r, canvas_t = transform_bounds(
                    src_crs, WGS84, src.bounds.left, src.bounds.bottom,
                    src.bounds.right, src.bounds.top)
            else:
                canvas_l, canvas_b, canvas_r, canvas_t = (
                    src.bounds.left, src.bounds.bottom, src.bounds.right, src.bounds.top)
        dst_t = _from_bounds(canvas_l, canvas_b, canvas_r, canvas_t, ow, oh)
        raw = np.full((oh, ow), np.nan, dtype=np.float32)
        _warp(source=rasterio.band(src, 1), destination=raw,
              src_transform=src.transform, src_crs=src_crs,
              dst_transform=dst_t, dst_crs=WGS84,
              resampling=Resampling.bilinear, src_nodata=nd, dst_nodata=np.nan)
        raw = raw.astype(float)
        if nd is not None: raw[raw == float(nd)] = np.nan

    goa_ext = [canvas_l, canvas_r, canvas_b, canvas_t]
    valid = raw[~np.isnan(raw)]
    vmin = float(np.nanpercentile(valid, 2))  if len(valid) > 0 else 0
    vmax = float(np.nanpercentile(valid, 98)) if len(valid) > 0 else 1

    v_range = vmax - vmin if vmax > vmin else 1
    thresholds = {
        "Low":       (vmin + v_range*0.2, vmin + v_range*0.4),
        "Moderate":  (vmin + v_range*0.4, vmin + v_range*0.6),
        "High":      (vmin + v_range*0.6, vmin + v_range*0.8),
        "Very High": (vmin + v_range*0.8, vmax),
    }

    import matplotlib.patheffects as pe

    fig, ax = plt.subplots(figsize=(9, 10), facecolor='#dff0ee')
    ax.set_facecolor('#dff0ee')

    if gdf is not None:
        for _, row in gdf.iterrows():
            geom = row.geometry
            gs = [geom] if geom.geom_type == 'Polygon' else list(geom.geoms)
            name = row[ncol] if ncol else ''
            is_selected = (name == target)
            for g in gs:
                xs, ys = g.exterior.xy
                if not is_selected:
                    ax.fill(list(xs), list(ys), color='#c8d8e4', alpha=0.85, zorder=1)
                    ax.plot(list(xs)+[xs[0]], list(ys)+[ys[0]],
                            color='#8fa8b8', linewidth=0.8, zorder=2, alpha=0.7)

    img = ax.imshow(raw, extent=goa_ext, cmap=RISK_CMAP, vmin=vmin, vmax=vmax,
                    origin='upper', aspect='auto', zorder=3, alpha=0.97, interpolation='bilinear')

    if selected_geom is not None:
        sel_gs = [selected_geom] if selected_geom.geom_type=='Polygon' else list(selected_geom.geoms)
        all_verts, all_codes = [], []
        for g in sel_gs:
            ring = list(g.exterior.coords) + [g.exterior.coords[0]]
            all_verts += [(c[0], c[1]) for c in ring]
            all_codes += [Path.MOVETO] + [Path.LINETO]*(len(ring)-2) + [Path.CLOSEPOLY]
        if all_verts:
            clip = PathPatch(Path(all_verts, all_codes),
                             transform=ax.transData, facecolor='none', edgecolor='none')
            ax.add_patch(clip); img.set_clip_path(clip)

        for g in sel_gs:
            xs, ys = g.exterior.xy
            shadow_xs = [x + 0.008 for x in xs]
            shadow_ys = [y - 0.010 for y in ys]
            ax.fill(shadow_xs+[shadow_xs[0]], shadow_ys+[shadow_ys[0]],
                    color='#0a2a26', alpha=0.35, zorder=2.5)
            ax.plot(list(xs)+[xs[0]], list(ys)+[ys[0]], color='white', linewidth=7.0, zorder=7, alpha=0.6)
            ax.plot(list(xs)+[xs[0]], list(ys)+[ys[0]], color='#a7f3d0', linewidth=4.5, zorder=8, alpha=0.7)
            ax.plot(list(xs)+[xs[0]], list(ys)+[ys[0]], color='#0d9488', linewidth=2.2, zorder=9)

        b = selected_geom.bounds
        ax.text((b[0]+b[2])/2, (b[1]+b[3])/2, taluk,
                ha='center', va='center', fontsize=12, fontweight='bold', color='#0f2027',
                bbox=dict(boxstyle='round,pad=0.45', facecolor='white',
                          edgecolor='#0d9488', alpha=0.96, linewidth=1.8),
                zorder=10, path_effects=[pe.withStroke(linewidth=3, foreground='white')])

    padx = (canvas_r - canvas_l) * 0.03; pady = (canvas_t - canvas_b) * 0.03
    ax.set_xlim(canvas_l - padx, canvas_r + padx)
    ax.set_ylim(canvas_b - pady, canvas_t + pady)
    ax.set_title(f"{taluk}  —  {MODEL_FULL_NAMES.get(model,model)}",
                 fontsize=11, pad=14, color='#0f2027', fontweight='bold')
    ax.tick_params(colors='#94a3b8', labelsize=7)
    for sp in ax.spines.values(): sp.set_edgecolor('#c8d8e4')
    ax.set_xlabel('Longitude', color='#94a3b8', fontsize=8)
    ax.set_ylabel('Latitude', color='#94a3b8', fontsize=8)
    cbar = plt.colorbar(img, ax=ax, fraction=0.028, pad=0.02, shrink=0.55)
    cbar.ax.tick_params(colors='#475569', labelsize=8)
    cbar.outline.set_edgecolor('#e2e8f0')
    cbar.set_label('Risk Level', color='#475569', fontsize=9)
    fig.patch.set_facecolor('#dff0ee'); plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=120, bbox_inches='tight', facecolor='#dff0ee')
    buf.seek(0); plt.close(fig)
    image_b64 = base64.b64encode(buf.getvalue()).decode()

    class_stats = {}
    total_valid = len(valid)
    for cls_name, (lo, hi) in thresholds.items():
        mask = (valid >= lo) & (valid <= hi)
        pct = round(100.0 * mask.sum() / total_valid, 1) if total_valid > 0 else 0
        class_stats[cls_name] = pct

    now_str=datetime.now().strftime("%d %b %Y, %H:%M")
    CLASS_COLOURS = {
        "Low": "#a6d96a", "Moderate": "#fee08b",
        "High": "#f46d43", "Very High": "#a50026",
    }
    class_bars = "".join(f"""
      <div style="display:flex;align-items:center;gap:10px;margin-bottom:12px">
        <div style="width:14px;height:14px;border-radius:3px;background:{CLASS_COLOURS[cls]};flex-shrink:0;border:1px solid rgba(0,0,0,.1)"></div>
        <div style="flex:1">
          <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:4px">
            <span style="font-family:'Space Mono',monospace;font-size:9px;letter-spacing:1px;color:var(--mid);text-transform:uppercase">{cls}</span>
            <span style="font-family:'Space Mono',monospace;font-size:9px;color:var(--teal);font-weight:700">{class_stats[cls]}%</span>
          </div>
          <div style="height:3px;border-radius:2px;background:rgba(13,148,136,.1);overflow:hidden">
            <div style="height:100%;width:{class_stats[cls]}%;background:{CLASS_COLOURS[cls]};border-radius:2px"></div>
          </div>
        </div>
      </div>""" for cls in ["Low","Moderate","High","Very High"])

    body=f"""
<div style="display:grid;grid-template-columns:1fr 290px;gap:24px;max-width:1100px;align-items:start">
  <div class="result-card" style="margin-top:0">
    <div class="result-header">
      <div>
        <div class="result-title">Risk Assessment Result</div>
        <div style="font-family:'Space Mono',monospace;font-size:10px;color:var(--teal);margin-top:6px;letter-spacing:1px">{MODEL_FULL_NAMES.get(model,model).split('—')[1].strip() if '—' in MODEL_FULL_NAMES.get(model,'') else model}</div>
        <div class="result-meta">{now_str} · {session['username']}</div>
      </div>
      <div class="badge-row">
        <span class="badge badge-teal">📍 {taluk}</span>
        <span class="badge badge-em">Risk Map</span>
      </div>
    </div>
    <img class="result-img" src="data:image/png;base64,{image_b64}" alt="Risk Map">
    <div class="result-footer">
      <span class="file-name">{filename}</span>
      <a href="/dashboard" class="back-link-btn">← New Assessment</a>
    </div>
  </div>
  <div style="display:flex;flex-direction:column;gap:16px;margin-top:0">
    <div style="background:rgba(255,255,255,.88);border:1.5px solid rgba(255,255,255,.5);border-radius:20px;padding:24px;box-shadow:0 8px 32px rgba(13,148,136,.08);backdrop-filter:blur(20px)">
      <div style="font-family:'Space Mono',monospace;font-size:9px;letter-spacing:2px;text-transform:uppercase;color:var(--soft);margin-bottom:20px">Risk Level Breakdown</div>
      {class_bars}
      <div style="margin-top:16px;padding-top:14px;border-top:1px solid var(--border);font-family:'Space Mono',monospace;font-size:8px;color:var(--soft);line-height:1.8;letter-spacing:0.5px">
        How much of {taluk} falls into each risk category, based on the {model} model.
      </div>
    </div>
    <div style="background:rgba(255,255,255,.88);border:1.5px solid rgba(255,255,255,.5);border-radius:20px;padding:24px;box-shadow:0 8px 32px rgba(13,148,136,.08);backdrop-filter:blur(20px)">
      <div style="font-family:'Space Mono',monospace;font-size:9px;letter-spacing:2px;text-transform:uppercase;color:var(--soft);margin-bottom:18px">Colour Guide</div>
      <div style="font-size:13px;color:var(--mid);line-height:2">
        <div style="display:flex;align-items:center;gap:8px;margin-bottom:4px"><span style="color:#a6d96a;font-size:16px">●</span><span><strong>Low / Moderate</strong> — Lower risk zones</span></div>
        <div style="display:flex;align-items:center;gap:8px"><span style="color:#a50026;font-size:16px">●</span><span><strong>High / Very High</strong> — Needs attention</span></div>
      </div>
    </div>
    <div style="background:rgba(13,148,136,.08);border:1.5px solid rgba(13,148,136,.2);border-radius:20px;padding:18px;backdrop-filter:blur(10px)">
      <div style="font-family:'Space Mono',monospace;font-size:8px;letter-spacing:2px;text-transform:uppercase;color:var(--soft);margin-bottom:10px">Model Used</div>
      <div style="font-family:'DM Serif Display',serif;font-size:16px;color:var(--ink)">{model}</div>
      <div style="font-family:'Space Mono',monospace;font-size:9px;color:var(--soft);margin-top:4px;letter-spacing:0.5px">{MODEL_FULL_NAMES.get(model,'').split('—')[1].strip() if '—' in MODEL_FULL_NAMES.get(model,'') else ''}</div>
    </div>
  </div>
</div>"""
    return render_layout(body,f"Risk Assessment Result",f"{taluk} · {model}","assess")

# =============================================================================
#  ADMIN ROUTES
# =============================================================================
@app.route('/admin')
@admin_required
def admin_dashboard():
    conn=get_db()
    try:
        total_users=conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        total_assessments=conn.execute("SELECT COUNT(*) FROM assessment_logs").fetchone()[0]
        unique_talukas=conn.execute("SELECT COUNT(DISTINCT taluk) FROM assessment_logs").fetchone()[0]
        recent=conn.execute("SELECT * FROM assessment_logs ORDER BY id DESC LIMIT 5").fetchall()
    finally:
        conn.close()
    rows="".join(
        f"<tr><td>{r['username']}</td><td>{r['taluk']}</td><td>{r['model']}</td>"
        f"<td style='font-size:11px;color:var(--soft)'>{r['timestamp'][:16]}</td></tr>"
        for r in recent
    ) or "<tr><td colspan='4' style='text-align:center;color:var(--soft);padding:20px;font-family:Space Mono,monospace;font-size:11px'>No assessments yet.</td></tr>"
    excel_s=(f"<span style='color:#15803d;font-family:Space Mono,monospace;font-size:10px;letter-spacing:1px'>✓ User records file active</span>"
             if EXCEL_OK and os.path.exists(EXCEL_PATH) else "<span style='color:#b91c1c;font-size:11px'>⚠ User records file not found</span>")
    gdf_s=(f"<span style='color:#15803d;font-family:Space Mono,monospace;font-size:10px;letter-spacing:1px'>✓ Area boundaries loaded ({len(get_gdf())} areas)</span>"
           if GEOPANDAS_OK and get_gdf() is not None else "<span style='color:#b91c1c;font-size:11px'>⚠ Area boundaries not found</span>")
    body=f"""
<div class="stats-grid">
  <div class="stat-card"><div class="stat-card-label">Registered Users</div><div class="stat-card-value">{total_users}</div></div>
  <div class="stat-card"><div class="stat-card-label">Total Assessments Run</div><div class="stat-card-value">{total_assessments}</div></div>
  <div class="stat-card"><div class="stat-card-label">Areas Assessed</div><div class="stat-card-value">{unique_talukas}</div></div>
</div>
<div style="margin-bottom:10px;padding:12px 18px;background:rgba(34,197,94,.06);border:1.5px solid rgba(34,197,94,.15);border-radius:10px">📊 {excel_s}</div>
<div style="margin-bottom:18px;padding:12px 18px;background:rgba(34,197,94,.06);border:1.5px solid rgba(34,197,94,.15);border-radius:10px">🗺️ {gdf_s}</div>
<div class="table-wrap">
  <div class="table-head">Recent Activity</div>
  <table><thead><tr><th>User</th><th>Area</th><th>Model</th><th>Date & Time</th></tr></thead>
  <tbody>{rows}</tbody></table>
</div>"""
    return render_layout(body,"Admin Dashboard","Overview of users and activity","admin")

@app.route('/admin/logs')
@admin_required
def admin_logs():
    conn=get_db()
    try:
        logs=conn.execute("SELECT * FROM assessment_logs ORDER BY id DESC").fetchall()
    finally:
        conn.close()
    rows="".join(
        f"<tr><td>{r['id']}</td><td>{r['username']}</td><td>{r['taluk']}</td>"
        f"<td>{r['model']}</td><td style='font-size:11px;color:var(--soft)'>{r['timestamp'][:16]}</td></tr>"
        for r in logs
    ) or "<tr><td colspan='5' style='text-align:center;color:var(--soft);padding:20px'>No activity yet.</td></tr>"
    body=f"""
<div class="table-wrap">
  <div class="table-head">All Activity ({len(logs)} records)</div>
  <table><thead><tr><th>#</th><th>User</th><th>Area</th><th>Model</th><th>Date & Time</th></tr></thead>
  <tbody>{rows}</tbody></table>
</div>"""
    return render_layout(body,"Activity Log","Full history of all assessments run","logs")

@app.route('/admin/users')
@admin_required
def admin_users():
    conn=get_db()
    try:
        users=conn.execute("SELECT * FROM users ORDER BY id DESC").fetchall()
    finally:
        conn.close()
    rows="".join(
        f"<tr><td>{u['id']}</td><td><strong>{u['username']}</strong></td><td>{u['email']}</td>"
        f"<td><span style='padding:3px 10px;border-radius:50px;font-family:Space Mono,monospace;font-size:8px;font-weight:600;letter-spacing:1px;"
        f"background:{'rgba(217,119,6,.08);color:#92400e;border:1px solid rgba(217,119,6,.15)' if u['role']=='admin' else 'rgba(13,148,136,.1);color:#0f766e;border:1px solid rgba(13,148,136,.2)'}'>"
        f"{u['role']}</span></td><td style='font-family:Space Mono,monospace;font-size:10px;color:var(--soft)'>{u['created_at'][:16]}</td></tr>"
        for u in users
    ) or "<tr><td colspan='5' style='text-align:center;color:var(--soft);padding:20px'>No users found.</td></tr>"
    body=f"""
<div class="table-wrap">
  <div class="table-head">All Users ({len(users)} registered)</div>
  <table><thead><tr><th>#</th><th>Username</th><th>Email</th><th>Role</th><th>Joined</th></tr></thead>
  <tbody>{rows}</tbody></table>
</div>"""
    return render_layout(body,"Users","All registered accounts","users")

# =============================================================================
if __name__ == '__main__':
    init_excel()
    init_db()
    os.makedirs(os.path.join(BASE_DIR, 'static', 'images'), exist_ok=True)
    img_path  = os.path.join(BASE_DIR, 'static', 'images', 'goa_map_bg.jpg')
    logo_path = os.path.join(BASE_DIR, 'static', 'images', 'logo.jpg')
    print(f"\n  ╔══════════════════════════════════════════════════════════╗")
    print(f"  ║  EVA SYSTEM — Goa Landslide Risk Assessment            ║")
    print(f"  ╚══════════════════════════════════════════════════════════╝")
    print(f"  URL      : http://127.0.0.1:5000")
    print(f"  Admin    : username=admin  password=admin123")
    print(f"  Map BG   : {'✓' if os.path.exists(img_path) else '✗ MISSING'} — {img_path}")
    print(f"  Logo     : {'✓' if os.path.exists(logo_path) else '✗ MISSING'} — {logo_path}")
    app.run(debug=True)